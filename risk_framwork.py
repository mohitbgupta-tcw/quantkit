"""
Created on Tue Mar  7 16:22:16 2023

@author: ABDARS
"""

import os
import warnings
import pandas as pd
import numpy as np
from openpyxl import load_workbook
# from scipy.stats.mstats import winsorize
# import matplotlib.pyplot as plt
import math
import re

warnings.filterwarnings("ignore")

#function for deleting sheets except for 'Sheet1' because we need at least one sheet in an excel file
def delete_unused_sheets(path):
    wb = load_workbook(path)
    sheet_names = wb.sheetnames
    list_of_sheets_to_keep = ['Sheet1']
    for i in sheet_names:
        if i not in list_of_sheets_to_keep:
            del wb[i]
    wb.save(path)
    wb.close()  
    
#function for writing to excel
def write_to_excel(path,sheet,table):
    ExcelWorkbook = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = ExcelWorkbook
    table.to_excel(writer, sheet_name = str(sheet))
    writer.save()
    ExcelWorkbook.close()
    
#Transform Qualitative Fields into Quantitative equivalent metrics
def Transform(df_Transform, df_Pf):
    for i in range(len(df_Transform['Field'])):
        df_Pf[df_Transform['Field'].iloc[i]] = df_Pf[df_Transform['Field'].iloc[i]].replace(df_Transform['OldValue'].iloc[i], df_Transform['NewValue'].iloc[i], regex = True)

#Remove special characters so that can properly perform replace and sub functions
def regexList(List):
    for i in range(len(List)):
        if isinstance(List[i], str):
            List[i] = re.sub('[^a-zA-Z0-9]+', '', str(List[i]))
    return(List)

#define function to strip data and factors for each selected analyst 
def analyst_factors(IndexTable, AnalystTable, analyst_name, risk_category, sector, sub_sector, identifier):
    df1 = AnalystTable.loc[(AnalystTable['Analyst'] == analyst_name) & (AnalystTable['Risk Category'] == risk_category) 
                            & (AnalystTable['Sector'] == sector) & (AnalystTable['Sub-Sector'] == sub_sector)]
    df2 = IndexTable.loc[IndexTable['ESRM Module']==sub_sector]
    df3 = df2[df2.columns.intersection(df1['Indicator Field Name'])]
    df3.insert(0, 'Region_EM', df2['Region_EM'])
    df3.insert(0, 'ISSUER_CNTRY_DOMICILE', df2['ISSUER_CNTRY_DOMICILE'])
    df3.insert(0, 'GICS_SUB_IND', df2['GICS_SUB_IND'])
    df3.insert(0, 'Identifier', df2[identifier])
    df3.insert(0, 'ISSUER_NAME', df2['ISSUER_NAME'])
    return df3 

def transition_score_calc(FileName, FilePathImport, FilePathExport, Q_Low, Q_High, High_Threshold, df_ISS_Data):
    #Load Portfolio Data 
    df_Portfolio = pd.read_excel(FilePathImport + FileName, sheet_name = 'Portfolio', engine = 'openpyxl')
    Sector_Code = pd.read_excel(FilePathImport + FileName, sheet_name = 'Sector_Code', engine = 'openpyxl')
    df_Portfolio = df_Portfolio.merge(Sector_Code, on = 'Portfolio', how = 'left')
    df_Portfolio['ISIN'].fillna('NoISIN', inplace=True) #if ISIN field is empty, then populate with NoISIN
    
    #Load MSCI Indicators 
    df_MSCI_Data = pd.read_excel(FilePathImport + FileName, sheet_name = 'MSCI Data', engine = 'openpyxl')
    
    #Merge Portfolio, MSCI, ISS
    cols_to_use = df_MSCI_Data.columns.difference(df_Portfolio.columns)
    df_PortfolioFactors = df_Portfolio.merge(df_MSCI_Data[cols_to_use], how = 'left', left_on = 'ISIN', right_on = 'Client_ID' ) 
    df_PortfolioFactors = df_PortfolioFactors.merge(df_ISS_Data[['Security ISIN', 'ClimateGHGReductionTargets', 'BrownExpTotalCapExSharePercent', 'GreenExpTotalCapExSharePercent']], how = 'left', 
                                      left_on = 'ISIN', right_on = 'Security ISIN')
    
    #Load in the GICS and BCLASS mapping and merge data to df_PortfolioFactors
    GICSMapping = pd.read_excel(FilePathImport + 'Materiality Factors by Sector.xlsx', sheet_name = 'GICS', engine = 'openpyxl')
    BCLASSMapping = pd.read_excel(FilePathImport + 'Materiality Factors by Sector.xlsx', sheet_name = 'BCLASS', engine = 'openpyxl')
    df_PortfolioFactors_GICS = df_PortfolioFactors.loc[df_PortfolioFactors['Sector_Code'] == 'GICS'].merge(GICSMapping[['Analyst', 'Industry', 'GICS_SUB_IND','Transition Risk Module']], on = 'GICS_SUB_IND', how ='left')
    df_PortfolioFactors_BCLASS =  df_PortfolioFactors.loc[df_PortfolioFactors['Sector_Code'] == 'BCLASS'].merge(BCLASSMapping[['Analyst', 'Industry','BCLASS_Level4','Transition Risk Module']], on = 'BCLASS_Level4', how ='left')
    df_PortfolioFactors = pd.concat([df_PortfolioFactors_GICS, df_PortfolioFactors_BCLASS]).reset_index(drop=True)
    
    #Calculation Carbon Intensity for Scope 1+2+3 (some sales values are 0 or negative)
    df_PortfolioFactors['Carbon Intensity (Scope 123)'] = df_PortfolioFactors['CARBON_EMISSIONS_SCOPE123']/df_PortfolioFactors.loc[df_PortfolioFactors['SALES_USD_RECENT'] > 0, 'SALES_USD_RECENT']
    
    #Adjustments:
    df_PortfolioFactors.replace('Not Collected', np.nan, inplace=True)
    df_PortfolioFactors['GreenExpTotalCapExSharePercent'] = df_PortfolioFactors['GreenExpTotalCapExSharePercent'].astype(float)
    df_PortfolioFactors['Industry'].fillna('NoIndustryAssigned', inplace=True)
    
    #Fill Carbon Intensity NA's with the median partitioned for 'Industry'. Previously, we needed to winsorize, but not using Z-Scored values anymore
    df_PortfolioFactors['Carbon Intensity (Scope 123)'] = df_PortfolioFactors['Carbon Intensity (Scope 123)'].fillna(df_PortfolioFactors.groupby('Industry')['Carbon Intensity (Scope 123)'].transform('median'))
    #df_PortfolioFactors['Carbon Intensity (Scope 123)'] = df_PortfolioFactors.groupby('Industry')['Carbon Intensity (Scope 123)'].transform(lambda row: winsorize(row, limits=[0,0.08]))
    
    #Identifiy High/Low Transition Risk for this without an Industry Mapped , and change Industry Name accordingly 
    df_PortfolioFactors.loc[(df_PortfolioFactors['Industry'] == 'NoIndustryAssigned' ) & (df_PortfolioFactors['Carbon Intensity (Scope 123)'] > High_Threshold),'Transition Risk Module'] = 'High'
    df_PortfolioFactors.loc[(df_PortfolioFactors['Industry'] == 'NoIndustryAssigned') & (df_PortfolioFactors['Carbon Intensity (Scope 123)'] <= High_Threshold),'Transition Risk Module'] = 'Low'
    df_PortfolioFactors.loc[(df_PortfolioFactors['Transition Risk Module'] == 'High') & (df_PortfolioFactors['Industry'] == 'NoIndustryAssigned'), 'Industry'] = 'NoIndustryAssignedHigh'
    df_PortfolioFactors.loc[(df_PortfolioFactors['Transition Risk Module'] == 'Low') & (df_PortfolioFactors['Industry'] == 'NoIndustryAssigned'), 'Industry'] = 'NoIndustryAssignedLow'
    
    #Define list of distinct industries
    Industry_Uniq = df_PortfolioFactors['Industry'].unique()
    
    
    #AddZ-Score Column per Industry field
    #df_PortfolioFactors.insert(len(df_PortfolioFactors.columns), 'Carbon Intensity Z_Score', df_PortfolioFactors.groupby(['Industry'])['Carbon Intensity (Scope 123)'].transform(lambda x : (x-x.mean()) / x.std()))
    
    #Add - Target Score or CapEx >30%
    def calc_new_col(row):
        if row['HAS_COMMITTED_TO_SBTI_TARGET'] == 1: #msci
            return -1
        elif row['HAS_SBTI_APPROVED_TARGET'] == 1: #msci
            return -1
        elif row['ClimateGHGReductionTargets'] == 'Ambitious Target': #sdg
            return -1
        elif row['ClimateGHGReductionTargets'] == 'Approved SBT': #sdg
            return -1
        elif row['ClimateGHGReductionTargets'] == 'Committed SBT': #sdg
            return -1
        elif row['RENEW_ENERGY_CAPEX_VS_TOTAL_CAPEX_PCT'] >= 30: #msci
            return -1
        elif row['GreenExpTotalCapExSharePercent'] >= 0.30: #sdg
            return -1
        else:
            return 0
    df_PortfolioFactors['Target_Score'] = df_PortfolioFactors.apply(calc_new_col, axis=1)
    
    #Create Coverage Table 
    Coverage = {'Industry': [] , 'Coverage': [] }
    Coverage = pd.DataFrame(data = Coverage)
    
    #Loop through all industries to calculate Coverage
    for a in range(len(Industry_Uniq)):
       Coverage_Data = df_PortfolioFactors[df_PortfolioFactors['Industry'] == Industry_Uniq[a]].count()

       df = pd.DataFrame({'Industry': Industry_Uniq[a], 'Coverage': [Coverage_Data.iloc[2]/Coverage_Data.iloc[0]]})
       Coverage = pd.concat([Coverage, df])
    
    
    #Pull Initial Transition Risk Score by Sector (High or Low)
    Sector_Avg = df_PortfolioFactors[['Industry', 'Transition Risk Module']].drop_duplicates().dropna()
    Sector_Avg.reset_index(drop = True, inplace=True)

    #Z-Score Carbon Intensity
    #df_PortfolioFactors['Carbon Intensity Z_Score'] = df_PortfolioFactors['Carbon Intensity Z_Score'].fillna(df_PortfolioFactors.groupby(['Industry'])['Carbon Intensity Z_Score'].transform('mean'))
    
    #Define Initial Stage [set at 5/5 (high intensity) and 3/5 (low intensity]
    df_PortfolioFactors.insert(len(df_PortfolioFactors.columns),'Initial_Score',0)
    df_PortfolioFactors.insert(len(df_PortfolioFactors.columns),'Middle_Score',0)
    df_PortfolioFactors.insert(len(df_PortfolioFactors.columns),'Transition_Score',0)
    Quantiles = pd.DataFrame({'Industry': Sector_Avg['Industry'].tolist(), 'Q_Low' : 0.0, 'Q_High': 0.0})
    for a in range(len(Sector_Avg)):
        if Sector_Avg['Transition Risk Module'][a] == 'High':
            df_PortfolioFactors['Initial_Score'].loc[df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]] = 5
            Low = df_PortfolioFactors['Carbon Intensity (Scope 123)'].loc[df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]].quantile([Q_Low, Q_High])[Q_Low]
            High = df_PortfolioFactors['Carbon Intensity (Scope 123)'].loc[df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]].quantile([Q_Low, Q_High])[Q_High]
            Quantiles.loc[Quantiles['Industry']==Sector_Avg['Industry'][a],'Q_Low']= Low
            Quantiles.loc[Quantiles['Industry']==Sector_Avg['Industry'][a],'Q_High'] = High
            df_PortfolioFactors['Middle_Score'].loc[(df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]) & (df_PortfolioFactors['Carbon Intensity (Scope 123)'] < Low)] = 5 - 2 #initial score + quantile score
            df_PortfolioFactors['Middle_Score'].loc[(df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]) & (df_PortfolioFactors['Carbon Intensity (Scope 123)'] >= Low) & (df_PortfolioFactors['Carbon Intensity (Scope 123)'] < High)] = 5 - 1
            df_PortfolioFactors['Middle_Score'].loc[(df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]) & (df_PortfolioFactors['Carbon Intensity (Scope 123)'] >= High)] = 5 + 0
            
        else:
            df_PortfolioFactors['Initial_Score'].loc[df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]] = 3
            Low = df_PortfolioFactors['Carbon Intensity (Scope 123)'].loc[df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]].quantile([Q_Low, Q_High])[Q_Low]
            High = df_PortfolioFactors['Carbon Intensity (Scope 123)'].loc[df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]].quantile([Q_Low, Q_High])[Q_High]
            Quantiles.loc[Quantiles['Industry']==Sector_Avg['Industry'][a],'Q_Low']= Low
            Quantiles.loc[Quantiles['Industry']==Sector_Avg['Industry'][a],'Q_High'] = High
            df_PortfolioFactors['Middle_Score'].loc[(df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]) & (df_PortfolioFactors['Carbon Intensity (Scope 123)'] < Low)] = 3 - 2 #initial score + quantile score
            df_PortfolioFactors['Middle_Score'].loc[(df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]) & (df_PortfolioFactors['Carbon Intensity (Scope 123)'] >= Low) & (df_PortfolioFactors['Carbon Intensity (Scope 123)'] < High)] = 3 - 1
            df_PortfolioFactors['Middle_Score'].loc[(df_PortfolioFactors['Industry']==Sector_Avg['Industry'][a]) & (df_PortfolioFactors['Carbon Intensity (Scope 123)'] >= High)] = 3 + 0
    df_PortfolioFactors['Transition_Score'] = np.maximum(np.minimum(df_PortfolioFactors['Middle_Score'] + df_PortfolioFactors['Target_Score'],5),1)
    
    
    #Change GIC_BY column to Sector 
    Coverage.rename(columns = {'Industry': 'Transition_Sector'}, inplace=True)
    df_PortfolioFactors.rename(columns = {'Industry': 'Transition_Sector'}, inplace=True)
    Quantiles.rename(columns = {'Industry': 'Transition_Sector'}, inplace=True)
    
    delete_unused_sheets(FilePathExport + 'Transition_Factors.xlsx') 
    #write to excel Coverage and Carbon Intensity Scope
    write_to_excel(FilePathExport +'Transition_Factors.xlsx', 'Coverage', Coverage)
    write_to_excel(FilePathExport +'Transition_Factors.xlsx', 'Carbon_Intensity_Data', df_PortfolioFactors[df_PortfolioFactors['ISIN'] != 'NoISIN'])
    write_to_excel(FilePathExport +'Transition_Factors.xlsx', 'Quantiles', Quantiles)
    
    #Calculate ESRM and Governance Score
def ESRM_Gov_score_calc(FileName, FilePathImport, FilePathExport):
    #Define Your Path
    df_Analyst_Table = pd.read_excel(FilePathImport + 'Materiality Factors by Sector.xlsx', sheet_name = 'Materiality Map_v2', skiprows = [0,1,2], engine = 'openpyxl')
    df_Analyst_Sectors = df_Analyst_Table[['Analyst', 'Sector', 'Risk Category']].drop_duplicates().apply(tuple, axis=1)
    #df_Analyst_Sectors = [('Lily', 'Technology & Telecoms', 'ESRM')] #testing code
    #df_Analyst_Sectors = [('Britz', 'Governance', 'Governance')] #testing code
    
    #Load Portfolio Data   
    df_Portfolio = pd.read_excel(FilePathImport + FileName, sheet_name = 'Portfolio', engine = 'openpyxl')
    Sector_Code = pd.read_excel(FilePathImport + FileName, sheet_name = 'Sector_Code', engine = 'openpyxl')
    df_Portfolio = df_Portfolio.merge(Sector_Code, on = 'Portfolio', how = 'left')
    df_Portfolio['ISIN'].fillna('NoISIN', inplace=True)
    
    #Load MSCI Indicators 
    df_MSCI_Data = pd.read_excel(FilePathImport + FileName, sheet_name = 'MSCI Data', engine = 'openpyxl')
    
    #Merge df_Portfolio, MSCI
    cols_to_use = df_MSCI_Data.columns.difference(df_Portfolio.columns)
    df_Portfolio_MSCI = df_Portfolio.merge(df_MSCI_Data[cols_to_use], how = 'left', left_on = 'ISIN', right_on = 'Client_ID' ) 
    #Fill GICS NA to Unassigned GICS - that way we can still calculate ESRM score even if GICS is not available:
    df_Portfolio_MSCI['GICS_SUB_IND'].fillna('Unassigned GICS', inplace=True)
    Pf_List = df_Portfolio['Portfolio'].unique().tolist() 
    
    
    GICSMapping = pd.read_excel(FilePathImport + 'Materiality Factors by Sector.xlsx', sheet_name = 'GICS', engine = 'openpyxl')
    BCLASSMapping = pd.read_excel(FilePathImport + 'Materiality Factors by Sector.xlsx', sheet_name = 'BCLASS', engine = 'openpyxl')
    Portfolio_GICS = df_Portfolio_MSCI.loc[df_Portfolio_MSCI['Sector_Code'] == 'GICS'].merge(GICSMapping[['Analyst', 'Industry', 'GICS_SUB_IND','ESRM Module']], on = 'GICS_SUB_IND', how ='left')
    Portfolio_BCLASS =  df_Portfolio_MSCI.loc[df_Portfolio_MSCI['Sector_Code'] == 'BCLASS'].merge(BCLASSMapping[['Analyst', 'Industry','BCLASS_Level4','ESRM Module']], on = 'BCLASS_Level4', how ='left')
    df_Portfolio_MSCI = pd.concat([Portfolio_GICS, Portfolio_BCLASS]).reset_index(drop=True)
     
    df_Regions = pd.read_excel(FilePathImport + 'Materiality Factors by Sector.xlsx', sheet_name = 'Regions', keep_default_na = False, engine = 'openpyxl')

    
    Sector_Framework_PBI_ESRM = pd.DataFrame([])
    Sector_Framework_PBI_GOV = pd.DataFrame([])

    for i in Pf_List:
        print(i)
        #Setup File 
        TabName = i
        Identifier = 'ISIN'
     
        #Load Data
        #Portfolio Data
        df_Pf = df_Portfolio_MSCI[df_Portfolio_MSCI['Portfolio'] == i]
        
        #Add and filter by region 
        df_Pf = df_Pf.merge(df_Regions, left_on = 'ISSUER_CNTRY_DOMICILE', right_on ='ISO2')
        df_Pf = df_Pf.reset_index(drop=True)  
        
        for x in df_Analyst_Sectors:
            AnalystName  = x[0]
            Sector = x[1]
            RiskCategory = x[2]
            
            #pull list of analyst SubSectors excel file
            df_Analyst_SubSectors = df_Analyst_Table.loc[(df_Analyst_Table['Analyst'] == AnalystName) & (df_Analyst_Table['Risk Category'] == RiskCategory) &
                                                       (df_Analyst_Table['Sector'] == Sector), 'Sub-Sector'].unique()  
            #df_Analyst_SubSectors = df_Analyst_SubSectors['Sub-Sector'].unique() 

            #iterate through each of the SubSectors 
            for z in range(len(df_Analyst_SubSectors)):
                print('Processing ' + AnalystName + '_' + Sector + '_' + RiskCategory + '_' + df_Analyst_SubSectors[z] + '.')   
                
                #Load Index Data and Transformation Data (changing qualitative to quantitative metrics)
                df_Transform = pd.read_excel(FilePathImport + 'Materiality Factors by Sector.xlsx', sheet_name = 'Transformation', engine = 'openpyxl')
                  
  
                #go through the transformation file to remove special characters so that it can be replaced with numerical values 
                df_Transform['OldValue'] = regexList(df_Transform['OldValue'])
                for i in range(len(df_Pf[df_Transform['Field'].unique()].columns)):
                    df_Pf[df_Pf[df_Transform['Field'].unique()].columns[i]] = regexList(df_Pf[df_Transform['Field'].unique()].iloc[:,i])
                Transform(df_Transform, df_Pf)
                
                #Load in rest of the data, including Analyst Indicators and Crosswalk 
                df_AnalystFactors = pd.read_excel(FilePathImport + 'Materiality Factors by Sector.xlsx', sheet_name = 'Materiality Map_v2', skiprows = [0,1,2], engine = 'openpyxl')
                
                #Combine Analyst Factors to Industry Codes and Pull in the Weights for specific Sub-Sectors 
                df_AnalystFactors_Weights = df_AnalystFactors.loc[(df_AnalystFactors['Analyst'] == AnalystName) & (df_AnalystFactors['Risk Category'] == RiskCategory) &
                                                           (df_AnalystFactors['Sector'] == Sector) & (df_AnalystFactors['Sub-Sector'] == df_Analyst_SubSectors[z])]
            
                
                #if any of the analyst factors include an arithmetic property, create new field in the index table
                
                # for a in df_AnalystFactors_Weights['Indicator Field Name']:
                #     #reindex = df_AnalystFactors_Weights['Indicator Field Name'].reset_index(drop=True)
                #     if a.find('/') >= 1:
                #         partitioned_string = a.partition('/')
                #         df_Pf[a] = df_Pf[partitioned_string[0]]/df_Pf[partitioned_string[2]]
                #     if a.find('**') >= 1:
                #         partitioned_string = a.partition('**')
                #         df_Pf[a] = df_Pf[partitioned_string[0]]*df_Pf[partitioned_string[2]]
                #     df_Pf[a] = pd.to_numeric(df_Pf[a])
                 
                    
                if RiskCategory == 'Governance':
                    df_Pf['Industry'] = df_Pf['Region']
                    df_Pf['ESRM Module'] = df_Pf['Region']
                
                #define function to strip data and factors for each selected analyst 
                df_AnalystData = analyst_factors(df_Pf, df_AnalystFactors, AnalystName, RiskCategory, Sector, df_Analyst_SubSectors[z], Identifier)
                df_AnalystData.insert(0, 'Pf', TabName)
                df_AnalystData_NumericalCols = df_AnalystData.select_dtypes(include=[np.number]).columns

                
                df_AnalystData['Flagged'] = ''
                for i in df_AnalystData_NumericalCols:
                    if df_AnalystData.empty:
                            break
                    if df_AnalystFactors_Weights.loc[df_AnalystFactors_Weights['Indicator Field Name'] == i, 'Operator'].values == '>':
                        df_AnalystData.loc[(df_AnalystData[i] <= df_AnalystFactors_Weights.loc[df_AnalystFactors_Weights['Indicator Field Name'] == i, 'Flag_Threshold'].item()), i + '_Flag'] = 0
                        df_AnalystData.loc[(df_AnalystData[i] > df_AnalystFactors_Weights.loc[df_AnalystFactors_Weights['Indicator Field Name'] == i, 'Flag_Threshold'].item()) | df_AnalystData[i].isna(), i  + '_Flag'] = 1
                        df_AnalystData.loc[df_AnalystData[i].isna(), i  + '_isnan'] = 1
                    elif df_AnalystFactors_Weights.loc[df_AnalystFactors_Weights['Indicator Field Name'] == i, 'Operator'].values == '=':
                        df_AnalystData.loc[(df_AnalystData[i] != df_AnalystFactors_Weights.loc[df_AnalystFactors_Weights['Indicator Field Name'] == i, 'Flag_Threshold'].item()), i  + '_Flag'] = 0 
                        df_AnalystData.loc[(df_AnalystData[i] == df_AnalystFactors_Weights.loc[df_AnalystFactors_Weights['Indicator Field Name'] == i, 'Flag_Threshold'].item()) | df_AnalystData[i].isna(), i  + '_Flag'] = 1
                        df_AnalystData.loc[df_AnalystData[i].isna(), i  + '_isnan'] = 1
                    elif df_AnalystFactors_Weights.loc[df_AnalystFactors_Weights['Indicator Field Name'] == i, 'Operator'].values == '<':
                        df_AnalystData.loc[(df_AnalystData[i] >= df_AnalystFactors_Weights.loc[df_AnalystFactors_Weights['Indicator Field Name'] == i, 'Flag_Threshold'].item()), i  + '_Flag'] = 0 
                        df_AnalystData.loc[(df_AnalystData[i] < df_AnalystFactors_Weights.loc[df_AnalystFactors_Weights['Indicator Field Name'] == i, 'Flag_Threshold'].item()) | df_AnalystData[i].isna(), i  + '_Flag'] = 1
                        df_AnalystData.loc[df_AnalystData[i].isna(), i  + '_isnan'] = 1                      
                        
                
                if RiskCategory == 'ESRM': 
                    df_AnalystData['ESRM_Flagged'] = df_AnalystData.loc[:,df_AnalystData.columns.str.contains('_Flag')].sum(axis = 'columns') 
                    df_AnalystData['ESRM_IsNAN'] = df_AnalystData.loc[:,df_AnalystData.columns.str.contains('_isnan')].sum(axis = 'columns') 
                    df_AnalystData['ESRM_Score'] = ''
                    df_AnalystData.loc[(df_AnalystData['ESRM_Flagged'] >= df_AnalystFactors_Weights['Well-Performing'].max(axis=0)) & (df_AnalystData['Region_EM'] == 'DM'), 'ESRM_Score'] = 1
                    df_AnalystData.loc[(df_AnalystData['ESRM_Flagged'] >= df_AnalystFactors_Weights['Above Average'].max(axis=0)) & (df_AnalystData['Region_EM'] == 'DM'), 'ESRM_Score'] = 2
                    df_AnalystData.loc[(df_AnalystData['ESRM_Flagged'] >= df_AnalystFactors_Weights['Average'].max(axis=0)) & (df_AnalystData['Region_EM'] == 'DM'), 'ESRM_Score'] = 3
                    df_AnalystData.loc[(df_AnalystData['ESRM_Flagged'] >= df_AnalystFactors_Weights['Below Average'].max(axis=0)) & (df_AnalystData['Region_EM'] == 'DM'), 'ESRM_Score'] = 4
                    df_AnalystData.loc[(df_AnalystData['ESRM_Flagged'] >= df_AnalystFactors_Weights['Concerning'].max(axis=0)) & (df_AnalystData['Region_EM'] == 'DM'), 'ESRM_Score'] = 5
                    
                    df_AnalystData.loc[(df_AnalystData['ESRM_Flagged'] >= df_AnalystFactors_Weights['Well-Performing-EM'].max(axis=0)) & (df_AnalystData['Region_EM'] == 'EM'), 'ESRM_Score'] = 1
                    df_AnalystData.loc[(df_AnalystData['ESRM_Flagged'] >= df_AnalystFactors_Weights['Above Average-EM'].max(axis=0)) & (df_AnalystData['Region_EM'] == 'EM'), 'ESRM_Score'] = 2
                    df_AnalystData.loc[(df_AnalystData['ESRM_Flagged'] >= df_AnalystFactors_Weights['Average-EM'].max(axis=0)) & (df_AnalystData['Region_EM'] == 'EM'), 'ESRM_Score'] = 3
                    df_AnalystData.loc[(df_AnalystData['ESRM_Flagged'] >= df_AnalystFactors_Weights['Below Average-EM'].max(axis=0)) & (df_AnalystData['Region_EM'] == 'EM'), 'ESRM_Score'] = 4
                    df_AnalystData.loc[(df_AnalystData['ESRM_Flagged'] >= df_AnalystFactors_Weights['Concerning-EM'].max(axis=0)) & (df_AnalystData['Region_EM'] == 'EM'), 'ESRM_Score'] = 5
                    
                    #Add dataframe across for PowerBI
                    df_AnalystData.loc[(df_AnalystData['Identifier'] == 'NoISIN'), 'ESRM_Score'] = 4
                    Sector_Framework_PBI_ESRM = [Sector_Framework_PBI_ESRM, df_AnalystData]
                    Sector_Framework_PBI_ESRM = pd.concat(Sector_Framework_PBI_ESRM)
                    
                elif RiskCategory == 'Governance':      
                    df_AnalystData['Governance_Flagged'] = df_AnalystData.loc[:,df_AnalystData.columns.str.contains('_Flag')].sum(axis = 'columns') 
                    df_AnalystData['Governance_IsNAN'] = df_AnalystData.loc[:,df_AnalystData.columns.str.contains('_isnan')].sum(axis = 'columns') 
                    df_AnalystData['Governance_Score'] = ''
                    df_AnalystData.loc[df_AnalystData['Governance_Flagged'] >= df_AnalystFactors_Weights['Well-Performing'].max(axis=0), 'Governance_Score'] = 1
                    df_AnalystData.loc[df_AnalystData['Governance_Flagged'] >= df_AnalystFactors_Weights['Above Average'].max(axis=0), 'Governance_Score'] = 2
                    df_AnalystData.loc[df_AnalystData['Governance_Flagged'] >= df_AnalystFactors_Weights['Average'].max(axis=0), 'Governance_Score'] = 3
                    df_AnalystData.loc[df_AnalystData['Governance_Flagged'] >= df_AnalystFactors_Weights['Below Average'].max(axis=0), 'Governance_Score'] = 4
                    df_AnalystData.loc[df_AnalystData['Governance_Flagged'] >= df_AnalystFactors_Weights['Concerning'].max(axis=0), 'Governance_Score'] = 5         
                    
                    #Add dataframe across for PowerBI
                    df_AnalystData.loc[(df_AnalystData['Identifier'] == 'NoISIN'), 'Governance_Score'] = 4
                    Sector_Framework_PBI_GOV = [Sector_Framework_PBI_GOV, df_AnalystData]
                    Sector_Framework_PBI_GOV = pd.concat(Sector_Framework_PBI_GOV)
            
     
    os.chdir(FilePathExport)
    delete_unused_sheets(FilePathExport + 'Sector_Framework_PBI.xlsx')
    write_to_excel(FilePathExport + 'Sector_Framework_PBI.xlsx', 'Sector_Framework_ESRM', Sector_Framework_PBI_ESRM)
    write_to_excel(FilePathExport + 'Sector_Framework_PBI.xlsx', 'Sector_Framework_GOV', Sector_Framework_PBI_GOV)
    
#Caclulate Themes 
def themes_calc(FileName, FilePathImport, FilePathExport, df_ISS_Data_Themes):
    #Load Portfolio Data 
    df_Portfolio = pd.read_excel(FilePathImport + FileName, sheet_name = 'Portfolio', engine = 'openpyxl') #pull in all portfolio holdings
    df_Portfolio['ISIN'].fillna('NoISIN', inplace=True) #set empty ISIN's to "NOISIN"
    Fields = pd.read_excel(FilePathImport + FileName, sheet_name = 'Fields', engine = 'openpyxl') #pull in MSCI and ISS Field indicators to use
    SectorCodeTable = pd.read_excel(FilePathImport + FileName, sheet_name = 'Sector_Code', engine = 'openpyxl') #pull which sector code to use (gics or bclass) for each portfolio

    #Load MSCI data and structure ISS Data 
    df_MSCI_Data = pd.read_excel(FilePathImport + FileName, sheet_name = 'MSCI Data', engine = 'openpyxl')
    df_ISS_Data_Themes = df_ISS_Data_Themes[Fields['ISS'].dropna()]
    df_ISS_Data_Themes[Fields['ISS_Numeric'].dropna()] = df_ISS_Data_Themes[Fields['ISS_Numeric'].dropna()].apply(pd.to_numeric, errors = 'coerce')

    #Merge Portfolio, MSCI, ISS data w/ Security Mapping
    cols_to_use = df_MSCI_Data.columns.difference(df_Portfolio.columns)
    df_Portfolio_MSCI = df_Portfolio.merge(df_MSCI_Data[cols_to_use], how = 'left', left_on = 'ISIN', right_on = 'Client_ID' ) 
    df_Portfolio_ISS = df_Portfolio.merge(df_ISS_Data_Themes, how = 'left', left_on = 'ISIN', right_on = 'Security ISIN' )  
    df_Portfolio_ISS.rename(columns = {'ISIN_x': 'ISIN'}, inplace = True)
    Pf_List = df_Portfolio['Portfolio'].unique().tolist() 

    #Pull Article 8 and 9 exclusions from spreadsheet
    ExclusionsA9 = pd.read_excel(FilePathImport+ 'Exclusions.xlsx', sheet_name = 'Article9', engine = 'openpyxl')
    ExclusionsA8 = pd.read_excel(FilePathImport+ 'Exclusions.xlsx', sheet_name = 'Article8', engine = 'openpyxl')

    # Load in ESRM, Gov, and Transition Scores
    ESRM = pd.read_excel(FilePathExport + 'Sector_Framework_PBI.xlsx', sheet_name = 'Sector_Framework_ESRM', engine = 'openpyxl')
    GOV = pd.read_excel(FilePathExport + 'Sector_Framework_PBI.xlsx', sheet_name = 'Sector_Framework_GOV', engine = 'openpyxl')
    Transition = pd.read_excel(FilePathExport + 'Transition_Factors.xlsx', sheet_name = 'Carbon_Intensity_Data', engine = 'openpyxl')

    #Bring in the Region/Country Factors 
    df_RegionFactors = pd.read_excel(FilePathImport+ 'Materiality Factors by Sector.xlsx', sheet_name = 'Regions', keep_default_na = False, engine = 'openpyxl')


    df_Themes = pd.DataFrame([])
    for i in Pf_List:
        
        print(i)
        #Setup File
        if 'BCLASS' in SectorCodeTable.loc[SectorCodeTable['Portfolio'] == i, 'Sector_Code']:
            Sector_Code = 'BCLASS_LEVEL4'
        else:
            Sector_Code = 'GICS_SUB_IND'
        Identifier = 'ISIN'

        #Carve out Portfolio Data
        df_Portfolio_MSCI_i = df_Portfolio_MSCI[df_Portfolio_MSCI['Portfolio'] == i]
        df_Portfolio_MSCI_i = df_Portfolio_MSCI_i[Fields['MSCI'].dropna()]
        df_Portfolio_MSCI_i.insert(0, 'Pf', i) 
        
        df_Portfolio_ISS_i = df_Portfolio_ISS[df_Portfolio_ISS['Portfolio'] == i]
        df_Portfolio_ISS_i = df_Portfolio_ISS_i[Fields['ISS'].dropna()]    

       
        df_Portfolio_MSCI_i = df_Portfolio_MSCI_i.merge(df_RegionFactors, how = 'left', left_on = 'ISSUER_CNTRY_DOMICILE', right_on ='ISO2')
        
        #Sustainabile/Transition Criteria      
        ThematicMapping = pd.read_excel(FilePathImport+ 'Themes.xlsx', sheet_name = 'Thematic Mapping', engine = 'openpyxl')
        SustainableIndustryMapping = pd.read_excel(FilePathImport+ 'Themes.xlsx', sheet_name = 'Sustainable Industry Mapping', engine = 'openpyxl')
        TransitionIndustryMapping = pd.read_excel(FilePathImport+ 'Themes.xlsx', sheet_name = 'Transition Industry Mapping', engine = 'openpyxl')
        TransitionIndustryMapping = TransitionIndustryMapping[['Transition_Target', 'Transition_Revenue', 'Acronym', Sector_Code]]
        Analyst_Adjustments = pd.read_excel(FilePathImport+ 'Themes.xlsx', sheet_name = 'Analyst Adjustment', engine = 'openpyxl')
        SecuritizedMapping = pd.read_excel(FilePathImport+ 'Themes.xlsx', sheet_name = 'SecuritizedMapping', engine = 'openpyxl')
        
        #Combine ISS and MSCI Data into one table 
        cols_to_use = df_Portfolio_ISS_i.columns.difference(df_Portfolio_MSCI_i.columns).tolist()
        cols_to_use.append(Identifier) 
        df_DataThemes = pd.merge(df_Portfolio_MSCI_i, df_Portfolio_ISS_i[cols_to_use], how='left', on = Identifier)

        #Clean up any non-numerical values to 0
        df_DataThemes.replace('Not Collected', np.nan, inplace=True)
        df_DataThemes.fillna(0, inplace=True)

        #Start with Sustaianbility Tag Calculation - Load in ISS and MSCI fields for each of the Thematic Categories and make changes to some of the data points
        ThemeUniques = ThematicMapping['Acronym'].unique()
        for i in ThemeUniques:
            ThematicMappingSet = ThematicMapping[ThematicMapping['Acronym'] == i] #pull each thematic category (i.e renewable energy)
            ISS1Uniques = ThematicMappingSet['ISS 1'].unique().tolist() #pull each distinct factor from set ISS 1
            ISS2Uniques = ThematicMappingSet['ISS 2'].unique().tolist() #pull each distinct factor from set ISS 2
            ISS1Uniques.extend(ISS2Uniques) #combine ISS 1 and ISS2
            ISS1Uniques = [x for x in ISS1Uniques if str(x) != 'nan'] #if there are NA's, remove it from the list 
            MSCISubUniques = ThematicMappingSet['MSCI Subcategories'].unique() #pull each distinct factor from MSCI Sub=Category  
            MSCISubUniques = [x for x in MSCISubUniques if str(x) != 'nan'] #if there are NA's, remove it from the list 
            
            df_DataThemes[i+'_MSCI'] = df_DataThemes[MSCISubUniques].sum(axis=1) #for each thematic category, add up all the associated MSCI data
            df_DataThemes[i+'_ISS'] = df_DataThemes[ISS1Uniques].max(axis=1)*100  #for each thematic category, take the max of the identified ISS fields (multiply ISS by 100 to compare to MSCI)
            df_DataThemes[i+'_code'] = ''
            
            ISS1UniquesKey = list(map(lambda x: x.replace('Percent', 'Prod'),ISS1Uniques))
            df_DataThemes[i + '_ISSKeyAdd'] = np.nan 
            for x in ISS1UniquesKey:
                check_list = df_DataThemes[x]
                contains = [check_list.astype(str).str.contains(i) for i in [j for j in ThematicMappingSet['ProductKeyAdd'] if str(j) != 'nan']]

                import pdb
                pdb.set_trace()

                for x in contains:
                    x.fillna(False, inplace=True)            
                df_DataThemes.loc[df_DataThemes[i + '_ISSKeyAdd'].isnull(), i + '_ISSKeyAdd']  = check_list[np.any(contains, axis=0)] 
        
        #Correct for Renewenergy, mobility, and smarticities energy (ISS data is too broad). 
        #If the description does not include any of the keywords, then set the ISS revenue to 0 for the following themes:
        df_DataThemes.loc[df_DataThemes['SMARTCITIES_ISSKeyAdd'].isnull(), 'SMARTCITIES_ISS'] = 0
        df_DataThemes.loc[df_DataThemes['MOBILITY_ISSKeyAdd'].isnull(), 'MOBILITY_ISS'] = 0
        df_DataThemes.loc[df_DataThemes['RENEWENERGY_ISSKeyAdd'].isnull(), 'RENEWENERGY_ISS'] = 0
           
     
        #Calculate CapEx
        df_DataThemes['GreenExpTotalCapExSharePercent'] = df_DataThemes['GreenExpTotalCapExSharePercent']*100 #multiply ISS CapEx by 100 to compare to MSCI
        df_DataThemes['CapEx'] = df_DataThemes[['GreenExpTotalCapExSharePercent', 'RENEW_ENERGY_CAPEX_VS_TOTAL_CAPEX_PCT']].max(axis=1)  #CapEx data is only one field so need the max for ISS and MSCI


        #Apply Sector Additions and Deletions across Themes 
        RENEWENERGY_SECTOR_DEL = ['Application Software', 'Leisure Products', 'Advertising', 'Internet Services & Infrastructure', 'Interactive Media & Services', 'Internet Services & Infrastructure']
        NUTRITION_SECTOR_DEL = ['Soft Drinks']  
        BIODIVERSITY_SECTOR_DEL = ['Apparel, Accessories & Luxury Goods', 'Apparel Retail', 'Retailers']
        HEALTH_SECTOR_ADD1 = ['Pharmaceuticals', 'Biotechnology']
        HEALTH_SECTOR_ADD2 = ['Life Sciences Tools & Services', 'Health Care Technology', 'Health Care Equipment', 'Health Care Supplies', 'Health Care Services']
        HEALTH_SECTOR_ADD3 = ['Health Care Distributors', 'Health Care Facilities', 'Managed Health Care']
        ALL_SECTOR_DEL = ['Casinos & Gaming', 'Distillers & Vintners', 'Brewers', 'Tobacco', 'Gaming']


        df_DataThemes = df_DataThemes.astype({'PALM_TIE':'str'})
        
        #Sustainability List
        df_DataThemes['Sustainability_Tag']=''
        df_DataThemes['Sustainability_Category']=''
        
        df_DataThemes.loc[((~df_DataThemes['GICS_SUB_IND'].isin(RENEWENERGY_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(RENEWENERGY_SECTOR_DEL)) 
                               | ~df_DataThemes['RENEWENERGY_ISSKeyAdd'].isnull()) #Not in Select GICS Industry Codes
                          & ((df_DataThemes['RENEWENERGY_MSCI'] >= 20) 
                          | (df_DataThemes['RENEWENERGY_ISS'] >= 20) 
                          | ((df_DataThemes[['RENEWENERGY_MSCI','RENEWENERGY_ISS']].max(axis=1) >= 15) 
                             & (df_DataThemes['CapEx'] >= 30)))
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL),'RENEWENERGY_code'] = 'RENEWENERGY'
        
        df_DataThemes.loc[((df_DataThemes['MOBILITY_MSCI'] >= 20) 
                              | ((df_DataThemes['MOBILITY_ISS'] >= 20) 
                                   & (df_DataThemes['GICS_SUB_IND'].isin(SustainableIndustryMapping.loc[SustainableIndustryMapping['Acronym'] == 'MOBILITY', 'GICS_SUB_IND'])
                                        | ~df_DataThemes['MOBILITY_ISSKeyAdd'].isnull()))
                                | ((df_DataThemes[['MOBILITY_MSCI','MOBILITY_ISS']].max(axis=1) >= 15) 
                             & (df_DataThemes['CapEx'] >= 30)))
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL),'MOBILITY_code'] = 'MOBILITY'  
        
        
        df_DataThemes.loc[((df_DataThemes['CIRCULARITY_MSCI'] >= 20) 
                           | ((df_DataThemes['CIRCULARITY_ISS'] >= 20) 
                              & (df_DataThemes['GICS_SUB_IND'].isin(SustainableIndustryMapping.loc[SustainableIndustryMapping['Acronym'] == 'CIRCULARITY', 'GICS_SUB_IND'])
                                 | ~df_DataThemes['CIRCULARITY_ISSKeyAdd'].isnull())))
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL), 'CIRCULARITY_code'] = 'CIRCULARITY'   

        df_DataThemes.loc[((df_DataThemes['CCADAPT_MSCI'] >= 20) 
                          | (df_DataThemes['CCADAPT_ISS'] >= 20))
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL), 'CCADAPT_code'] = 'CCADAPT'     

        df_DataThemes.loc[~df_DataThemes['GICS_SUB_IND'].isin(BIODIVERSITY_SECTOR_DEL) 
                          & ~df_DataThemes['BCLASS_Level4'].isin(BIODIVERSITY_SECTOR_DEL)
                          & ~(df_DataThemes['PALM_TIE'] == 'True')  
                          & ((df_DataThemes['BIODIVERSITY_MSCI'] >= 20) 
                             | (df_DataThemes['BIODIVERSITY_ISS'] >= 20))
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL), 'BIODIVERSITY_code'] = 'BIODIVERSITY'   
            
        df_DataThemes.loc[((df_DataThemes['SMARTCITIES_MSCI'] >= 20) 
                           | ((df_DataThemes['SMARTCITIES_ISS'] >= 20) 
                              & (df_DataThemes['GICS_SUB_IND'].isin(SustainableIndustryMapping.loc[SustainableIndustryMapping['Acronym'] == 'SMARTCITIES', 'GICS_SUB_IND'])
                                  | ~df_DataThemes['SMARTCITIES_ISSKeyAdd'].isnull())))
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL), 'SMARTCITIES_code'] = 'SMARTCITIES'   

        df_DataThemes.loc[((df_DataThemes['EDU_MSCI'] >= 20) 
                               | (df_DataThemes['EDU_ISS'] >= 20))
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL), 'EDU_code'] = 'EDU'   

        df_DataThemes.loc[(((df_DataThemes['HEALTH_MSCI'] >= 20) & (((df_DataThemes['ACCESS_TO_HLTHCRE_SCORE'] >= 5) | (df_DataThemes['TRAILING_12M_RD_SALES'] >= 18)) & (df_DataThemes['GICS_SUB_IND'].isin(HEALTH_SECTOR_ADD1) | df_DataThemes['BCLASS_Level4'].isin(HEALTH_SECTOR_ADD1))))
                               | ((df_DataThemes['HEALTH_MSCI'] >= 20) & (df_DataThemes['GICS_SUB_IND'].isin(HEALTH_SECTOR_ADD2) | df_DataThemes['BCLASS_Level4'].isin(HEALTH_SECTOR_ADD2)))
                               | (df_DataThemes['GICS_SUB_IND'].isin(HEALTH_SECTOR_ADD3) | df_DataThemes['BCLASS_Level4'].isin(HEALTH_SECTOR_ADD3))
                               | (df_DataThemes['SI_ORPHAN_DRUGS_REV'] > 100)) 
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL), 'HEALTH_code'] = 'HEALTH'   
        
        df_DataThemes.loc[((df_DataThemes['SANITATION_MSCI'] >= 20) 
                              | (df_DataThemes['SANITATION_ISS'] >= 20))
                            & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL), 'SANITATION_code'] = 'SANITATION'   

        df_DataThemes.loc[((df_DataThemes['SI_SOCIAL_FIN__MAX_REV'] >= 20) 
                          | (df_DataThemes['SI_CONNECTIVITY_MAX_REV'] >= 5)
                          | ((df_DataThemes['INCLUSION_ISS'] >= 20) 
                             & (df_DataThemes['GICS_SUB_IND'].isin(SustainableIndustryMapping.loc[SustainableIndustryMapping['Acronym'] == 'INCLUSION', 'GICS_SUB_IND'])
                                | ~df_DataThemes['INCLUSION_ISSKeyAdd'].isnull()))) 
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL),'INCLUSION_code'] = 'INCLUSION'   
        
        df_DataThemes.loc[(~df_DataThemes['GICS_SUB_IND'].isin(NUTRITION_SECTOR_DEL) | ~df_DataThemes['BCLASS_Level4'].isin(NUTRITION_SECTOR_DEL))  #Not in Select GICS Industry Codes
                          & ((df_DataThemes['NUTRITION_MSCI'] >= 20) 
                             | (df_DataThemes['NUTRITION_ISS'] >= 20))
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL), 'NUTRITION_code'] = 'NUTRITION'   
        
        df_DataThemes.loc[((df_DataThemes['GICS_SUB_IND'].isin(['Health Care Factilities', 'Health Care REITs']) 
                                & (df_DataThemes['SI_BASIC_N_TOTAL_MAX_REV'] >= 20))  
                              | (df_DataThemes['AFFORDABLE_MSCI'] >= 20))
                          & ~df_DataThemes['GICS_SUB_IND'].isin(ALL_SECTOR_DEL) & ~df_DataThemes['BCLASS_Level4'].isin(ALL_SECTOR_DEL), 'AFFORDABLE_code'] = 'AFFORDABLE'   
        
            
                
        df_DataThemes['Primary_Rev_Sustainable'] = df_DataThemes[['AFFORDABLE_MSCI', 'AFFORDABLE_ISS', 
                                                            'BIODIVERSITY_MSCI', 'BIODIVERSITY_ISS', 
                                                            'CCADAPT_MSCI', 'CCADAPT_ISS', 
                                                            'CIRCULARITY_MSCI', 'CIRCULARITY_ISS', 
                                                            'EDU_MSCI', 'EDU_ISS', 
                                                            'HEALTH_MSCI', 'HEALTH_ISS', 
                                                            'INCLUSION_MSCI', 'INCLUSION_ISS', 
                                                            'MOBILITY_MSCI', 'MOBILITY_ISS', 
                                                            'NUTRITION_MSCI', 'NUTRITION_ISS', 
                                                            'RENEWENERGY_MSCI', 'RENEWENERGY_ISS', 
                                                            'SANITATION_MSCI', 'SANITATION_ISS', 
                                                            'SMARTCITIES_MSCI', 'SMARTCITIES_ISS']].idxmax(axis=1).str.split('_').apply(lambda x: x[0])

        #Create a Single Column Tag for Sustainability
        df_DataThemes.loc[((df_DataThemes['RENEWENERGY_code'] == 'RENEWENERGY') | (df_DataThemes['MOBILITY_code'] == 'MOBILITY') | (df_DataThemes['CIRCULARITY_code'] == 'CIRCULARITY') 
                     | (df_DataThemes['CCADAPT_code'] == 'CCADAPT') | (df_DataThemes['BIODIVERSITY_code'] == 'BIODIVERSITY') | (df_DataThemes['SMARTCITIES_code'] == 'SMARTCITIES') 
                     | (df_DataThemes['EDU_code'] == 'EDU') | (df_DataThemes['HEALTH_code'] == 'HEALTH') | (df_DataThemes['SANITATION_code'] == 'SANITATION') 
                     | (df_DataThemes['INCLUSION_code'] == 'INCLUSION') | (df_DataThemes['NUTRITION_code'] == 'NUTRITION')| (df_DataThemes['AFFORDABLE_code'] == 'AFFORDABLE')), 'Sustainability_Tag'] = 'Y'
        
        #Create a Single Column for listing out the name of the Categories 
        df_DataThemes['Sustainability_Category'] = df_DataThemes['RENEWENERGY_code'] + ' ' + df_DataThemes['MOBILITY_code'] + ' ' + df_DataThemes['CIRCULARITY_code'] + ' ' + df_DataThemes['CCADAPT_code'] + ' ' + df_DataThemes['BIODIVERSITY_code'] + ' ' + df_DataThemes['SMARTCITIES_code'] + ' ' + df_DataThemes['EDU_code'] + ' ' + df_DataThemes['HEALTH_code'] + ' ' + df_DataThemes['SANITATION_code'] + ' ' + df_DataThemes['INCLUSION_code'] + ' ' + df_DataThemes['NUTRITION_code'] + ' ' + df_DataThemes['AFFORDABLE_code']
        
        

        #################Calculate Transition Tag - process data first##################################
        df_DataThemes['SDGSolClimatePercentCombCont'] = df_DataThemes['SDGSolClimatePercentCombCont']*100 #Convert ISS to MSCI equivalent by multiplying 100 (we already have the CapEx)
        df_DataThemes['Climate_Revenue'] = df_DataThemes[['SDGSolClimatePercentCombCont', 'CT_CC_TOTAL_MAX_REV']].max(axis=1)  
        
        #Set Different Targets 
        Target_A = ['Approved SBT']
        Target_AA = ['Approved SBT','Ambitious Target']
        Target_AAC = ['Approved SBT','Committed SBT','Ambitious Target']
        Target_AACN = ['Approved SBT','Committed SBT','Ambitious Target','Non-Ambitious Target']
        Target_AC = ['Approved SBT','Committed SBT']
        Target_CA = ['Committed SBT','Ambitious Target']
        Target_CN = ['Committed SBT', 'Non-Ambitious Target']
        Target_N = ['Non-Ambitious Target']
     

        df_DataThemes['Transition_Tag']=''
        df_DataThemes['Transition_Category']=' '


        #Transition Tag   
        #OilandGas, CoalFuels, Indgas, and Utilities are distinct code 
        df_DataThemes.loc[(df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'OilandGas',Sector_Code]) 
                        & (((df_DataThemes['ClimateGHGReductionTargets'].isin(Target_AAC)) | (df_DataThemes['HAS_COMMITTED_TO_SBTI_TARGET'] == 1)  | (df_DataThemes['HAS_SBTI_APPROVED_TARGET'] == 1))
                        & ((df_DataThemes['CapEx']> 0) | (df_DataThemes['Climate_Revenue'] > 0)))),'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[(df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'OilandGas',Sector_Code]) 
                        & (((df_DataThemes['CapEx'] >= 15) & (df_DataThemes['Climate_Revenue'] > 0)) | ((df_DataThemes['CapEx']> 0) & (df_DataThemes['Climate_Revenue'] > 15)))),'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[(df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'OilandGas',Sector_Code]) 
                        & ((df_DataThemes['Climate_Revenue'] >= 25) | (df_DataThemes['CT_ALT_ENERGY_BIOFUEL_MAX_REV'] >25))),'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[(df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'CoalFuels',Sector_Code]))
                       & (((df_DataThemes['CT_ALT_ENERGY_MAX_REV'] >= 95) | (df_DataThemes['Climate_Revenue'] >= 95)) | (df_DataThemes['THERMAL_COAL_MAX_REV_PCT'] == 0)),'Transition_Tag'] = 'Y'  
        
        df_DataThemes.loc[(df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'IndGases',Sector_Code])) #missing Lithium involvement
                       & ((df_DataThemes['ClimateGHGReductionTargets'].isin(Target_AAC)) | (df_DataThemes['HAS_COMMITTED_TO_SBTI_TARGET'] == 1)  | (df_DataThemes['HAS_SBTI_APPROVED_TARGET'] == 1) 
                       | (df_DataThemes['ISSUER_NAME'].str.contains('Lithium'))),'Transition_Tag'] = 'Y'   #& df_DataThemes['ISSUER_CNTRY_DOMICILE'].isin(['US','UK','AU','NZ','CA']))),'Transition_Tag'] = 'Y'  
        
        df_DataThemes.loc[(df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'IndGases',Sector_Code]))
                       & (((df_DataThemes['ClimateGHGReductionTargets'].isin(Target_AACN)) | (df_DataThemes['HAS_COMMITTED_TO_SBTI_TARGET'] == 1)  | (df_DataThemes['HAS_SBTI_APPROVED_TARGET'] == 1)) 
                       & (df_DataThemes['Climate_Revenue'] >= 5)),'Transition_Tag'] = 'Y'      
            
        df_DataThemes.loc[(df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Utilities',Sector_Code]) 
                        & ((df_DataThemes['ClimateGHGReductionTargets'] == 'Approved SBT') | (df_DataThemes['HAS_SBTI_APPROVED_TARGET'] == 1))),'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[(df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Utilities',Sector_Code]) 
                        & ((df_DataThemes['ClimateGHGReductionTargets'] == 'Ambitious Target') & ((df_DataThemes['CapEx'] >= 30) | (df_DataThemes['Climate_Revenue'] >= 20)))),'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[(df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Utilities',Sector_Code]) 
                        & ((df_DataThemes['ClimateGHGReductionTargets'] == 'Committed SBT') | (df_DataThemes['HAS_COMMITTED_TO_SBTI_TARGET'] == 1)) & (df_DataThemes['Climate_Revenue'] >= 35)),'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[(df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Utilities',Sector_Code]) 
                        & (df_DataThemes['Climate_Revenue'] >= 40)),'Transition_Tag'] = 'Y'
        
        
        
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Target_A' ,Sector_Code])  
                        & (df_DataThemes['ClimateGHGReductionTargets'].isin(Target_A)
                           | (df_DataThemes['HAS_SBTI_APPROVED_TARGET'] == 1)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Target_AA' ,Sector_Code])  
                        & (df_DataThemes['ClimateGHGReductionTargets'].isin(Target_AA)
                           | (df_DataThemes['HAS_SBTI_APPROVED_TARGET'] == 1)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Target_AAC' ,Sector_Code])  
                        & (df_DataThemes['ClimateGHGReductionTargets'].isin(Target_AAC)
                           | (df_DataThemes['HAS_SBTI_APPROVED_TARGET'] == 1) | (df_DataThemes['HAS_COMMITTED_TO_SBTI_TARGET'] == 1)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Target_AACN' ,Sector_Code])  
                        & (df_DataThemes['ClimateGHGReductionTargets'].isin(Target_AACN)
                           | (df_DataThemes['HAS_SBTI_APPROVED_TARGET'] == 1) | (df_DataThemes['HAS_COMMITTED_TO_SBTI_TARGET'] == 1)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Target_AC' ,Sector_Code])  
                        & (df_DataThemes['ClimateGHGReductionTargets'].isin(Target_AC)
                           | (df_DataThemes['HAS_SBTI_APPROVED_TARGET'] == 1) | (df_DataThemes['HAS_COMMITTED_TO_SBTI_TARGET'] == 1)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Target_CA' ,Sector_Code])  
                        & (df_DataThemes['ClimateGHGReductionTargets'].isin(Target_CA)
                           | (df_DataThemes['HAS_COMMITTED_TO_SBTI_TARGET'] == 1)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Target_CN' ,Sector_Code])  
                        & (df_DataThemes['ClimateGHGReductionTargets'].isin(Target_CN)
                           | (df_DataThemes['HAS_COMMITTED_TO_SBTI_TARGET'] == 1)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Target_N' ,Sector_Code])  
                        & (df_DataThemes['ClimateGHGReductionTargets'].isin(Target_N)), 'Transition_Tag'] = 'Y'
        
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Revenue'] == 'Revenue_10' ,Sector_Code])  
                        & ((df_DataThemes['Climate_Revenue'] >= 10) | (df_DataThemes['CapEx']> 10)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Revenue'] == 'Revenue_20' ,Sector_Code])  
                        & ((df_DataThemes['Climate_Revenue'] >= 20) | (df_DataThemes['CapEx']> 10)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Revenue'] == 'Revenue_30' ,Sector_Code])  
                        & ((df_DataThemes['Climate_Revenue'] >= 30) | (df_DataThemes['CapEx']> 10)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Revenue'] == 'Revenue_40' ,Sector_Code])  
                        & ((df_DataThemes['Climate_Revenue'] >= 40) | (df_DataThemes['CapEx']> 10)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Revenue'] == 'Revenue_50' ,Sector_Code])  
                        & ((df_DataThemes['Climate_Revenue'] >= 50) | (df_DataThemes['CapEx']> 10)), 'Transition_Tag'] = 'Y'
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Revenue'] == 'Revenue_60' ,Sector_Code])  
                        & ((df_DataThemes['Climate_Revenue'] >= 60) | (df_DataThemes['CapEx']> 10)), 'Transition_Tag'] = 'Y'
        
        
        df_DataThemes.loc[df_DataThemes[Sector_Code].isin(TransitionIndustryMapping.loc[TransitionIndustryMapping['Transition_Target'] == 'Target_NRev' ,Sector_Code])  
                        & ((df_DataThemes['ClimateGHGReductionTargets'].isin(Target_N) & df_DataThemes['Climate_Revenue'] > 0) | (df_DataThemes['CapEx']> 10)), 'Transition_Tag'] = 'Y'
        
      
        #Map the Transition Categories to the Data via GICS/Sector Code 
        df_DataThemes = pd.merge(df_DataThemes, TransitionIndustryMapping[['Acronym', Sector_Code]], how='left', left_on = Sector_Code, right_on = Sector_Code) 
        df_DataThemes.loc[(df_DataThemes['Transition_Tag'] == 'Y'), 'Transition_Category'] = df_DataThemes['Acronym']
        
        #Remove Transition Names in specified GICS if they alreay have a Sustainability Tag 
        Remove_Transition_Set = ['Industrial Machinery', 'Motorcycle Manufacturers', 'Highways & Railtracks', 'Industrial Conglomerates', 'Electrical Components & Equipment', 'Heavy Electrical Equipment']
        df_DataThemes.loc[(df_DataThemes['GICS_SUB_IND'].isin(Remove_Transition_Set)) & (df_DataThemes['Sustainability_Tag'] == 'Y'), 'Transition_Category'] = None 
        df_DataThemes.loc[(df_DataThemes['GICS_SUB_IND'].isin(Remove_Transition_Set)) & (df_DataThemes['Sustainability_Tag'] == 'Y'), 'Transition_Tag'] = None 
        df_DataThemes.drop(['Acronym'], axis=1) #Drop this unused column 

        #Create column with either a Transition or Sustainability Tag
        #df_DataThemes.loc[(df_DataThemes['Transition_Tag'] == 'Y') | (df_DataThemes['Sustainability_Tag'] == 'Y'), 'SIG_Coverage'] = 'Y' 
        
        #Merge all the Risk Scores (GOV, ESRM, Transition) to the dataframe
        df_DataThemes = pd.merge(df_DataThemes, GOV[['Pf','Identifier', 'Governance_IsNAN', 'Governance_Flagged', 'Governance_Score', 
                                             'CARBON_EMISSIONS_CDP_DISCLOSURE_Flag', 'COMBINED_CEO_CHAIR_Flag', 'CONTROLLING_SHAREHOLDER_Flag', 'CROSS_SHAREHOLDINGS_Flag', 'FEMALE_DIRECTORS_PCT_Flag',
                                             'INDEPENDENT_BOARD_MAJORITY_Flag', 'NEGATIVE_DIRECTOR_VOTES_Flag', 'PAY_LINKED_TO_SUSTAINABILITY_Flag', 'POISON_PILL_Flag' ,'WOMEN_EXEC_MGMT_PCT_RECENT_Flag']], how='left', left_on = ['Pf',Identifier], right_on = ['Pf','Identifier']) 
        df_DataThemes = pd.merge(df_DataThemes, ESRM[['Pf','Identifier', 'ESRM_IsNAN', 'ESRM_Flagged', 'ESRM_Score',
                                              'BIODIV_LAND_USE_EXP_SCORE_Flag','COMM_REL_RISK_EXP_SCORE_Flag', 'CONTR_SUPPLY_CHAIN_LABOR_N_TOTAL_Flag','ENERGY_EFFICIENCY_EXP_SCORE_Flag',
                                              'LABOR_MGMT_EXP_SCORE_Flag','PRIVACY_DATA_SEC_EXP_SCORE_Flag','WATER_STRESS_EXP_SCORE_Flag','CUSTOMER_RELATIONS_SCORE_Flag','PROD_SFTY_QUALITY_EXP_SCORE_Flag']], how='left', left_on = ['Pf',Identifier], right_on = ['Pf','Identifier']) 
        df_DataThemes = pd.merge(df_DataThemes, Transition[['Portfolio', Identifier, 'Transition_Sector', 'Transition_Score','Carbon Intensity (Scope 123)','ClimateGHGReductionTargets','GreenExpTotalCapExSharePercent', 'RENEW_ENERGY_CAPEX_VS_TOTAL_CAPEX_PCT',
                                                'HAS_COMMITTED_TO_SBTI_TARGET','HAS_SBTI_APPROVED_TARGET']], how='left', left_on = ['Pf',Identifier], right_on = ['Portfolio',Identifier])
        
        
        #For non-applicable securities, apply a score of 0 across all risk scores 
        Sector_Level2 = ['Cash and Other', 'Muni / Local Authority', 'Sovereign', 'Residential MBS', 'CMBS', 'ABS']
        df_DataThemes.loc[(df_DataThemes['BCLASS_Level4'] == 'Treasury') | (df_DataThemes['ISIN'] == 'NoISIN') | (df_DataThemes['Sector Level 2'].isin(Sector_Level2)) | (df_DataThemes['ISSUER_NAME'].str.contains('TCW FUNDS')) | (df_DataThemes['ISSUER_NAME'].str.contains(' ETF ')),'Governance_Score'] = 0
        df_DataThemes.loc[(df_DataThemes['BCLASS_Level4'] == 'Treasury') | (df_DataThemes['ISIN'] == 'NoISIN') | (df_DataThemes['Sector Level 2'].isin(Sector_Level2)) | (df_DataThemes['ISSUER_NAME'].str.contains('TCW FUNDS')) | (df_DataThemes['ISSUER_NAME'].str.contains(' ETF ')),'ESRM_Score'] = 0
        df_DataThemes.loc[(df_DataThemes['BCLASS_Level4'] == 'Treasury') | (df_DataThemes['ISIN'] == 'NoISIN') | (df_DataThemes['Sector Level 2'].isin(Sector_Level2)) | (df_DataThemes['ISSUER_NAME'].str.contains('TCW FUNDS')) | (df_DataThemes['ISSUER_NAME'].str.contains(' ETF ')),'Transition_Score'] = 0
        
        #Calculate Score for Securitized Produts
        df_DataThemes['Securitized_Score'] = 0
        Sector_Level2 = ['Residential MBS', 'CMBS', 'ABS']
        df_DataThemes.loc[(df_DataThemes['Sector Level 2'].isin(Sector_Level2)) & ((df_DataThemes['Labeled ESG Type'] != 'None') | (df_DataThemes['Issuer ESG'] == 'Yes')) & (df_DataThemes['TCW ESG'] != 'None'),'Securitized_Score'] = 1
        Collat_Type = ['LEED Platinum', 'LEED Gold','LEED Silver','LEED Certified','LEED (Multi Property)','BREEAM Very Good']
        df_DataThemes.loc[(df_DataThemes['Sector Level 2'].isin(Sector_Level2)) & (df_DataThemes['ESG Collateral Type'].isin(Collat_Type)) &(df_DataThemes['Labeled ESG Type'] != 'Labeled Green') & (df_DataThemes['TCW ESG'] == 'TCW Green'),'Securitized_Score'] = 2
        Collat_Type = ['TCW Criteria','Small Business Loan','FFELP Student Loan']
        df_DataThemes.loc[(df_DataThemes['Sector Level 2'].isin(Sector_Level2)) & (df_DataThemes['ESG Collateral Type'].isin(Collat_Type)) & ~(df_DataThemes['ISSUER_NAME'].str.contains('TBA', na=False)) & (df_DataThemes['Labeled ESG Type'] != 'Labeled Social') & (df_DataThemes['TCW ESG'] == 'TCW Social'),'Securitized_Score'] = 2
        df_DataThemes.loc[(df_DataThemes['Sector Level 2'].isin(Sector_Level2)) & (df_DataThemes['ESG Collateral Type'] == 'ESG CLO'), 'Securitized_Score'] = 2
        df_DataThemes.loc[(df_DataThemes['Sector Level 2'].isin(Sector_Level2)) & (df_DataThemes['ISSUER_NAME'].str.contains('TBA')),'Securitized_Score'] = 3
        df_DataThemes.loc[(df_DataThemes['Sector Level 2'].isin(Sector_Level2)) & (df_DataThemes['Labeled ESG Type'] == 'None') & (df_DataThemes['TCW ESG'] == 'None') & ~(df_DataThemes['ISSUER_NAME'].str.contains('TBA', na=False)),'Securitized_Score'] = 4
        df_DataThemes.loc[(df_DataThemes['Sector Level 2'].isin(Sector_Level2)) & ((df_DataThemes['Labeled ESG Type'] != 'None') | (df_DataThemes['Issuer ESG'] == 'Yes')) & (df_DataThemes['TCW ESG'] == 'None'),'Securitized_Score'] = 5

        #Apply 0 sovereign score for non-sovereigns
        df_DataThemes.loc[~(df_DataThemes['BCLASS_Level4'] == 'Treasury') & ~(df_DataThemes['Sector Level 2'] == 'Sovereign'), 'Sovereign_Score'] = 0
        
        
        
        df_DataThemes = df_DataThemes.loc[:,~df_DataThemes.columns.duplicated()]
        #Include Analyst Adjustments
        df_DataThemes['Review_Flag'] =' '
        df_DataThemes['Review_Comments'] =' '
        df_DataThemes['Muni_Score'] = 0
        Analyst_Adjustments.sort_values(by = ['Adjustment'], ascending = False, inplace = True) #run through deletions first
        for index, row in Analyst_Adjustments.iterrows():
            if (row['Adjustment'] == 'Deletion') & ((row['Thematic Type'] == 'People') | (row['Thematic Type'] == 'Planet')):
                df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Sustainability_Tag'] = 'X*'
                df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Sustainability_Category'] = df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Sustainability_Category'].replace(row['Category'], '', regex=True) 

            if (row['Adjustment'] == 'Addition') & ((row['Thematic Type'] == 'People') | (row['Thematic Type'] == 'Planet')):
                df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Sustainability_Tag'] = 'Y*'
                df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Sustainability_Category'] = df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Sustainability_Category'] + ' ' + row['Category'] 
            
            if (row['Adjustment'] == 'Deletion') & (row['Thematic Type'] == 'Transition'):
                df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Transition_Tag'] = 'X*'
                df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Transition_Category'] = df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Transition_Category'].replace(row['Category'], '', regex=True) 
            
            if (row['Adjustment'] == 'Addition') & (row['Thematic Type'] == 'Transition'):
                df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Transition_Tag'] = 'Y*'
                df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Transition_Category']= df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Transition_Category'] + ' ' + row['Category']  
            
            if (row['Category'] == 'Governance') or (row['Category'] == 'Transition') or (row['Category'] == 'ESRM'): #using MSCI ISSUERID to reference security...since all the data is coming MSCI 
                if row['Adjustment'][0:5] == 'Score':
                    df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), row['Category']+'_Score'] = int(row['Adjustment'][-1])
                    df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Review_Flag'] = 'Analyst Adjustment'
                
                if row['Adjustment'] == 'No Change':
                    df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Review_Flag'] = 'Analyst Adjustment - May Require Action'
                
                if str(df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Review_Comments']) is not None:
                    df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Review_Comments'] += ' & ' + str(row['Comments'])
        
            #Assign Muni Score - no systematic approach, manual review by analyst
            if (row['Category'] == 'Muni'):
                if row['Adjustment'][0:5] == 'Score':
                    df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), row['Category']+'_Score'] = int(row['Adjustment'][-1])
                    df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Review_Flag'] = 'Analyst Adjustment'
                if str(df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Review_Comments']) is not None:
                    df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row[Identifier]) | (df_DataThemes[Identifier] == row[Identifier]), 'Review_Comments'] += ' & ' + str(row['Comments'])
        
        df_Themes  = [df_Themes , df_DataThemes]
        df_Themes  = pd.concat(df_Themes)  
        # loop done
        

    #map through security master issues
    df_SecurityMaster = df_MSCI_Data[['ISSUER_NAME','Client_ID', 'ISSUERID','PARENT_ISSUERID','PARENT_ULTIMATE_ISSUERID']] 
    df_DataThemes = df_Themes.drop_duplicates(subset=['Pf', Identifier, 'Portfolio_Weight'], keep = 'first')

    for index, row in df_SecurityMaster.iterrows():
        Sector_Level2 = ['Cash and Other', 'Muni / Local Authority', 'Sovereign', 'Residential MBS', 'CMBS', 'ABS']
        if (~df_DataThemes.loc[(df_DataThemes['ISIN'] == row['Client_ID']), 'Sector Level 2'].isin(Sector_Level2).any()) & (~df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['PARENT_ULTIMATE_ISSUERID']), 'Sector Level 2'].isin(Sector_Level2).any()):
            if ((df_DataThemes.loc[(df_DataThemes['ISIN'] == row['Client_ID']), 'ESRM_IsNAN'].values >= 4).any() | pd.isna(df_DataThemes.loc[(df_DataThemes['ISIN'] == row['Client_ID']), 'ESRM_IsNAN']).any()) & ~pd.isna(df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['PARENT_ULTIMATE_ISSUERID']), 'ESRM_Score'].max()) & ~pd.isna(row['PARENT_ULTIMATE_ISSUERID']):
                df_DataThemes.loc[df_DataThemes['ISIN'] == row['Client_ID'], 'ESRM_Score'] = df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['PARENT_ULTIMATE_ISSUERID']), 'ESRM_Score'].min()
                df_DataThemes.loc[df_DataThemes['ISIN'] == row['Client_ID'], 'Review_Comments'] += ' & ' + 'ESRM Mapped to Ultimate Parent'
            if ((df_DataThemes.loc[(df_DataThemes['ISIN'] == row['Client_ID']), 'Governance_IsNAN'].values >= 4).any() | pd.isna(df_DataThemes.loc[(df_DataThemes['ISIN'] == row['Client_ID']), 'Governance_IsNAN']).any()) & ~pd.isna(df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['PARENT_ULTIMATE_ISSUERID']), 'Governance_Score'].max()) & ~pd.isna(row['PARENT_ULTIMATE_ISSUERID']):
                df_DataThemes.loc[df_DataThemes['ISIN'] == row['Client_ID'], 'Governance_Score'] = df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['PARENT_ULTIMATE_ISSUERID']), 'Governance_Score'].min()
                df_DataThemes.loc[df_DataThemes['ISIN'] == row['Client_ID'], 'Review_Comments'] += ' & ' + 'Governance Mapped to Ultimate Parent'
        df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['ISSUERID']) | (df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'ESRM_Score'] = df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['ISSUERID']) | (df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'ESRM_Score'].min()
        df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['ISSUERID']) | (df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Governance_Score'] = df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['ISSUERID']) | (df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Governance_Score'].min()
        df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['ISSUERID']) | (df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Transition_Score'] = df_DataThemes.loc[(df_DataThemes['ISSUERID'] == row['ISSUERID']) | (df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Transition_Score'].min()
        df_DataThemes.loc[(df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Sustainability_Tag'] = df_DataThemes.loc[(df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Sustainability_Tag'].max()
        df_DataThemes.loc[(df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Transition_Tag'] = df_DataThemes.loc[(df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Transition_Tag'].max()
        df_DataThemes.loc[(df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Sustainability_Category'] = df_DataThemes.loc[(df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Sustainability_Category'].max()
        df_DataThemes.loc[(df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Transition_Category'] = df_DataThemes.loc[(df_DataThemes['ISSUER_NAME'] == row['ISSUER_NAME']), 'Transition_Category'].max()
        #Sector_Level2 = ['Residential MBS', 'CMBS', 'ABS']
        df_DataThemes.loc[~df_DataThemes['Sector Level 2'].isin(Sector_Level2) & (df_DataThemes['ISSUERID'] == row['ISSUERID']) & (row['ISSUERID'] != 0), 'ISSUER_NAME'] = df_DataThemes.loc[~df_DataThemes['Sector Level 2'].isin(Sector_Level2) & (df_DataThemes['ISSUERID'] == row['ISSUERID']) & (row['ISSUERID'] != 0), 'ISSUER_NAME'].max()

    
    #Set Risk Scores NAs to 5's and assign for review 
    df_DataThemes.loc[pd.isna(df_DataThemes['Governance_Score']),'Governance_Score'] = 5
    df_DataThemes.loc[pd.isna(df_DataThemes['ESRM_Score']),'ESRM_Score'] = 5
    df_DataThemes.loc[pd.isna(df_DataThemes['Transition_Score']),'Transition_Score'] = 5
        
    df_DataThemes.loc[(df_DataThemes['Governance_Score'] == 5), 'Review_Flag'] = 'Needs Review'
    df_DataThemes.loc[(df_DataThemes['ESRM_Score'] == 5),'Review_Flag'] = 'Needs Review'
    df_DataThemes.loc[(df_DataThemes['Transition_Score'] == 5),'Review_Flag'] = 'Needs Review'   
   
    #Define SIG Category 
    #df_DataThemes.loc[(df_DataThemes['Transition_Tag'] == 'Y*') | (df_DataThemes['Sustainability_Tag'] == 'Y*'), 'SIG_Coverage'] = 'Y'       
    #df_DataThemes.loc[((df_DataThemes['Transition_Tag'] == 'X*') & (df_DataThemes['Sustainability_Tag'] == '')) | ((df_DataThemes['Sustainability_Tag'] == 'X*') & (df_DataThemes['Transition_Tag'] == '')) , 'SIG_Coverage'] = '' 
                 
    df_DataThemes.loc[df_DataThemes['Transition_Category'].isnull(),'Transition_Category'] = ' '
    df_DataThemes['SClass_Level1']='Eligible'
    df_DataThemes['SClass_Level2']='ESG Scores'
    df_DataThemes['SClass_Level3']='ESG Scores'
    df_DataThemes['SClass_Level4']='Average ESG Score'
    df_DataThemes.loc[((df_DataThemes['Governance_Score'] + df_DataThemes['ESRM_Score'] + df_DataThemes['Transition_Score']) <= 6) & ((df_DataThemes['Governance_Score'] + df_DataThemes['ESRM_Score'] + df_DataThemes['Transition_Score']) > 0), 'SClass_Level4']='Leading ESG Score'  
    df_DataThemes.loc[(df_DataThemes['Sovereign_Score'] <= 2) & (df_DataThemes['Sovereign_Score'] > 0), 'SClass_Level4']='Leading ESG Score'   
    df_DataThemes.loc[(df_DataThemes['Securitized_Score'] <= 2) & (df_DataThemes['Securitized_Score'] > 0), 'SClass_Level4']='Leading ESG Score'  
    df_DataThemes.loc[(df_DataThemes['Muni_Score'] <= 2) & (df_DataThemes['Muni_Score'] > 0), 'SClass_Level4']='Leading ESG Score'  
    df_DataThemes.loc[(df_DataThemes['Governance_Score'] == 0) & (df_DataThemes['ESRM_Score'] == 0) & (df_DataThemes['Transition_Score'] == 0) & (df_DataThemes['Securitized_Score'] == 0) & (df_DataThemes['Sovereign_Score'] == 0) & (df_DataThemes['Muni_Score'] == 0), 'SClass_Level4'] = 'Not Scored'
    df_DataThemes['SClass_Level4-P'] = np.nan
    df_DataThemes['SClass_Level5']= df_DataThemes['ESG Collateral Type']
    
    
    #Hard Adjustments
    df_DataThemes.loc[df_DataThemes['ISSUER_NAME'].str.startswith('INTELSAT ', na=False), 'SClass_Level1'] = 'Preferred'
    df_DataThemes.loc[df_DataThemes['ISSUER_NAME'].str.startswith('INTELSAT ', na=False), 'SClass_Level2'] = 'Sustainable Theme'
    df_DataThemes.loc[df_DataThemes['ISSUER_NAME'].str.startswith('INTELSAT ', na=False), 'SClass_Level3'] = 'People'
    df_DataThemes.loc[df_DataThemes['ISSUER_NAME'].str.startswith('INTELSAT ', na=False), 'SClass_Level4'] = 'INCLUSION'
    df_DataThemes.loc[df_DataThemes['ISSUER_NAME'].str.startswith('INTELSAT ', na=False), 'SClass_Level4-P'] = 'INCLUSION'
    df_DataThemes.loc[df_DataThemes['ISSUER_NAME'].str.startswith('INTELSAT ', na=False), 'Governance_Score'] = 4 
    df_DataThemes.loc[df_DataThemes['ISSUER_NAME'].str.startswith('INTELSAT ', na=False), 'ESRM_Score'] = 1
    df_DataThemes.loc[df_DataThemes['ISSUER_NAME'].str.startswith('INTELSAT ', na=False), 'Review_Flag'] = ' '
    df_DataThemes.loc[df_DataThemes['ISSUER_NAME'].str.startswith('INTELSAT ', na=False), 'Review_Comments'] += '& Hard Adjustment & Data Scarcity - Some investor rights and oversight issues remain (controlling shareholder and non-independent board), but considering recent reemergence from Ch11 in 2022 may also have improver potential'
    
    #Assign A8 and A9 funds/benchmarks in order to apply exclusions respective 
    A9Funds = [6781,16705,3738,3750,16719,16706, 'BB HIGH YIELD 2% CAP', 'MSCI WORLD', 'RUSSELL 1000 GROWTH', 'RUSSELL 1000 VALUE', 'RUSSELL 3000 GROWTH INDEX', 'BB INTERMEDIATE CORP', 'JPM CEMBI BROAD DIVERSE']
    A8Funds = [13727,3735,3775,13751, 'BLOOMBERG AGGREGATE INDEX']
    
    for index, row in df_DataThemes.iterrows():
        Planet_Count = 0
        People_Count = 0
        Transition_Count = 0
        #SClass_Level4 = ''
        if 'LOWCARBON' in row['Transition_Category']:
            Transition_Count = Transition_Count+1
            SClass_Level4 = 'LOWCARBON'

        if 'CARBONACCOUNT' in row['Transition_Category']:
            Transition_Count = Transition_Count+1
            SClass_Level4 = 'CARBONACCOUNT'

        if 'PIVOTTRANSPORT' in row['Transition_Category']:
            Transition_Count = Transition_Count+1
            SClass_Level4 = 'PIVOTTRANSPORT'

        if 'REALASSETS' in row['Transition_Category']:
            Transition_Count = Transition_Count+1
            SClass_Level4 = 'REALASSETS'

        if 'MATERIALS' in row['Transition_Category']:
            Transition_Count = Transition_Count+1
            SClass_Level4 = 'MATERIALS'

        if 'AGRIFORESTRY' in row['Transition_Category']:
            Transition_Count = Transition_Count+1
            SClass_Level4 = 'AGRIFORESTRY'
        
        if 'HEALTH' in row['Sustainability_Category']:
            People_Count = People_Count+1
            SClass_Level4 = 'HEALTH'

        if 'SANITATION' in row['Sustainability_Category']:
            People_Count = People_Count+1
            SClass_Level4 = 'SANITATION'

        if 'EDU' in row['Sustainability_Category']:
            People_Count = People_Count+1
            SClass_Level4 = 'EDU'

        if 'INCLUSION' in row['Sustainability_Category']:
            People_Count = People_Count+1
            SClass_Level4 = 'INCLUSION'

        if 'NUTRITION' in row['Sustainability_Category']:
            People_Count = People_Count+1
            SClass_Level4 = 'NUTRITION'

        if 'AFFORDABLE' in row['Sustainability_Category']:
            People_Count = People_Count+1
            SClass_Level4 = 'AFFORDABLE'
            
        if 'RENEWENERGY' in row['Sustainability_Category']:
            Planet_Count = Planet_Count+1
            SClass_Level4 = 'RENEWENERGY'
            
        if 'MOBILITY' in row['Sustainability_Category']:
            Planet_Count = Planet_Count+1
            SClass_Level4 = 'MOBILITY'
            
        if 'CIRCULARITY' in row['Sustainability_Category']:
            Planet_Count = Planet_Count+1
            SClass_Level4 = 'CIRCULARITY'
            
        if 'CCADAPT' in row['Sustainability_Category']:
            Planet_Count = Planet_Count+1
            SClass_Level4 = 'CCADAPT'
            
        if 'BIODIVERSITY' in row['Sustainability_Category']:
            Planet_Count = Planet_Count+1
            SClass_Level4 = 'BIODIVERSITY'
            
        if 'SMARTCITIES' in row['Sustainability_Category']:
            Planet_Count = Planet_Count+1
            SClass_Level4 = 'SMARTCITIES'
 
            
        if (row['Transition_Tag'] =='Y' or row['Transition_Tag'] =='Y*') and (Transition_Count > 1):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Multi-Thematic'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Transition'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Transition'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Eligible'
        if (row['Transition_Tag'] =='Y' or row['Transition_Tag'] =='Y*')  and (Transition_Count == 1):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = SClass_Level4
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Transition'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Transition'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Eligible'
            
            
        if (' TBA ' in str(row['ISSUER_NAME'])):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'AFFORDABLE'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'People'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Sustainable Theme'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'
        
        
        if (row['Sustainability_Tag'] =='Y' or row['Sustainability_Tag'] =='Y*') and (Planet_Count > 1):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = row['Primary_Rev_Sustainable']
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Planet'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Multi-Thematic'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Sustainable Theme'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'
        if (row['Sustainability_Tag'] =='Y' or row['Sustainability_Tag'] =='Y*') and (People_Count > 1):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = row['Primary_Rev_Sustainable']
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'People'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Multi-Thematic'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Sustainable Theme'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'
        if (row['Sustainability_Tag'] =='Y' or row['Sustainability_Tag'] =='Y*') and (Planet_Count == 1):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = SClass_Level4
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Planet'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Sustainable Theme'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'
        if (row['Sustainability_Tag'] =='Y' or row['Sustainability_Tag'] =='Y*') and (People_Count == 1):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = SClass_Level4
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'People'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Sustainable Theme'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'
        if (row['Sustainability_Tag'] =='Y' or row['Sustainability_Tag'] =='Y*') and (People_Count > 0) and (Planet_Count > 0):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = row['Primary_Rev_Sustainable']
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Planet & People'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Multi-Thematic'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Sustainable Theme'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'
        
        if (row['ESG Collateral Type'] in (SecuritizedMapping['ESG Collat Type'].values)):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = SecuritizedMapping.loc[SecuritizedMapping['ESG Collat Type']==row['ESG Collateral Type'], 'Primary'].values[0]
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = SecuritizedMapping.loc[SecuritizedMapping['ESG Collat Type']==row['ESG Collateral Type'], 'Primary'].values[0]
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = SecuritizedMapping.loc[SecuritizedMapping['ESG Collat Type']==row['ESG Collateral Type'], 'Sclass_Level3'].values[0]
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Sustainable Theme'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'
        
        if (row['Sovereign_Score'] == 5):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = 'Poor Sovereign Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Poor Sovereign Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Excluded' 
        
        if (row['Labeled ESG Type'] == 'Labeled Green'):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = 'Green'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Green'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'ESG-Labeled Bonds'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'ESG-Labeled Bonds'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'
        
        if (row['Labeled ESG Type'] == 'Labeled Social'):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = 'Social'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Social'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'ESG-Labeled Bonds'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'ESG-Labeled Bonds'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'
            
        if (row['Labeled ESG Type'] == 'Labeled Sustainable'):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = 'Sustainable'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Sustainable'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'ESG-Labeled Bonds'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'ESG-Labeled Bonds'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'
        
        if (row['Labeled ESG Type'] == 'Labeled Sustainable Linked'):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = 'Sustainability-Linked Bonds'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Sustainability-Linked Bonds'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'ESG-Labeled Bonds'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'ESG-Labeled Bonds'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Preferred'        
            
        if (row['Transition_Score'] == 5):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = 'Poor Transition Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Poor Transition Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Excluded'
              
        if (row['ESRM_Score'] == 5) or (row['UNGC_COMPLIANCE'] == 'Fail') or (row['OVERALL_FLAG'] =='Red') or (row['IVA_COMPANY_RATING'] == 'CCC'):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = 'Poor ESRM Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Poor ESRM Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Excluded'     
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'ESRM_Score'] = 5
        
        if (row['Governance_Score'] == 5):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = 'Poor Governance Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Poor Governance Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Excluded' 
            
        if (row['Securitized_Score'] == 5):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = 'Poor Securitized Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Poor Securitized Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Excluded' 
        
        if (row['Muni_Score'] == 5):
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4-P'] = 'Poor Muni Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level4'] = 'Poor Muni Score'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level3'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level2'] = 'Exclusion'
            df_DataThemes.loc[df_DataThemes[Identifier] == row[Identifier], 'SClass_Level1'] = 'Excluded' 
        
        
         #Loop done 

    #Carve out for Labeled Bonds - Energy and Utility will be okay to hold if labeled)    
    Labeled_ESG_Type = ['Labeled Green', 'Labeled Social', 'Labeled Sustainable', 'Labeled Sustainable Linked']
    Sector_Carve_OUT = ['Electric Utilities', 'Gas Utilities','Multi-Utilities','Other Utility','Water Utilities','Independent Power Producers & Energy Traders','Integrated Oil & Gas','Oil & Gas Drilling','Oil & Gas Equipment & Services','Oil & Gas Exploration & Production','Oil & Gas Refining & Marketing','Oil & Gas Storage & Transportation','Oil Field Services']
    df_DataThemes.loc[(df_DataThemes['ISSUERID'].isin(ExclusionsA9['MSCI Issuer ID'])) & (df_DataThemes['Pf'].isin(A9Funds)) & ~df_DataThemes['Labeled ESG Type'].isin(Labeled_ESG_Type) & ~df_DataThemes['BCLASS_Level4'].isin(Sector_Carve_OUT), 'SClass_Level4-P'] = 'Excluded Sector'
    df_DataThemes.loc[(df_DataThemes['ISSUERID'].isin(ExclusionsA9['MSCI Issuer ID'])) & (df_DataThemes['Pf'].isin(A9Funds)) & ~df_DataThemes['Labeled ESG Type'].isin(Labeled_ESG_Type) & ~df_DataThemes['BCLASS_Level4'].isin(Sector_Carve_OUT), 'SClass_Level4'] = 'Excluded Sector'
    df_DataThemes.loc[(df_DataThemes['ISSUERID'].isin(ExclusionsA9['MSCI Issuer ID'])) & (df_DataThemes['Pf'].isin(A9Funds)) & ~df_DataThemes['Labeled ESG Type'].isin(Labeled_ESG_Type) & ~df_DataThemes['BCLASS_Level4'].isin(Sector_Carve_OUT), 'SClass_Level3'] = 'Exclusion'
    df_DataThemes.loc[(df_DataThemes['ISSUERID'].isin(ExclusionsA9['MSCI Issuer ID'])) & (df_DataThemes['Pf'].isin(A9Funds)) & ~df_DataThemes['Labeled ESG Type'].isin(Labeled_ESG_Type) & ~df_DataThemes['BCLASS_Level4'].isin(Sector_Carve_OUT), 'SClass_Level2'] = 'Exclusion'
    df_DataThemes.loc[(df_DataThemes['ISSUERID'].isin(ExclusionsA9['MSCI Issuer ID'])) & (df_DataThemes['Pf'].isin(A9Funds)) & ~df_DataThemes['Labeled ESG Type'].isin(Labeled_ESG_Type) & ~df_DataThemes['BCLASS_Level4'].isin(Sector_Carve_OUT), 'SClass_Level1'] = 'Excluded' 
        
    df_DataThemes.loc[(df_DataThemes['ISSUERID'].isin(ExclusionsA8['MSCI Issuer ID'])) & (df_DataThemes['Pf'].isin(A8Funds)) & ~df_DataThemes['Labeled ESG Type'].isin(Labeled_ESG_Type) & ~df_DataThemes['BCLASS_Level4'].isin(Sector_Carve_OUT), 'SClass_Level4-P'] = 'Excluded Sector'
    df_DataThemes.loc[(df_DataThemes['ISSUERID'].isin(ExclusionsA8['MSCI Issuer ID'])) & (df_DataThemes['Pf'].isin(A8Funds)) & ~df_DataThemes['Labeled ESG Type'].isin(Labeled_ESG_Type) & ~df_DataThemes['BCLASS_Level4'].isin(Sector_Carve_OUT), 'SClass_Level4'] = 'Excluded Sector'
    df_DataThemes.loc[(df_DataThemes['ISSUERID'].isin(ExclusionsA8['MSCI Issuer ID'])) & (df_DataThemes['Pf'].isin(A8Funds)) & ~df_DataThemes['Labeled ESG Type'].isin(Labeled_ESG_Type) & ~df_DataThemes['BCLASS_Level4'].isin(Sector_Carve_OUT), 'SClass_Level3'] = 'Exclusion'
    df_DataThemes.loc[(df_DataThemes['ISSUERID'].isin(ExclusionsA8['MSCI Issuer ID'])) & (df_DataThemes['Pf'].isin(A8Funds)) & ~df_DataThemes['Labeled ESG Type'].isin(Labeled_ESG_Type) & ~df_DataThemes['BCLASS_Level4'].isin(Sector_Carve_OUT), 'SClass_Level2'] = 'Exclusion'
    df_DataThemes.loc[(df_DataThemes['ISSUERID'].isin(ExclusionsA8['MSCI Issuer ID'])) & (df_DataThemes['Pf'].isin(A8Funds)) & ~df_DataThemes['Labeled ESG Type'].isin(Labeled_ESG_Type) & ~df_DataThemes['BCLASS_Level4'].isin(Sector_Carve_OUT), 'SClass_Level1'] = 'Excluded'   
            
    df_DataThemes['SClass_Level4-P'] = df_DataThemes['SClass_Level4-P'].fillna(df_DataThemes['SClass_Level4'])
    
    #Portfolio Contribution Score 
    df_DataThemes = df_DataThemes.reset_index()
    df_DataThemes['Portfolio_Score_Contribution'] = 0
    df_DataThemes['Corporate_Score'] = (df_DataThemes['Governance_Score'] + df_DataThemes['ESRM_Score'] + df_DataThemes['Transition_Score'])/3
    df_DataThemes['Portfolio_Score_Contribution'] = (df_DataThemes['Corporate_Score'] + df_DataThemes['Securitized_Score'] + df_DataThemes['Sovereign_Score']) * df_DataThemes['Portfolio_Weight'] 
    
    df_DataThemes['Portfolio_Score']  = (df_DataThemes['Corporate_Score'] + df_DataThemes['Securitized_Score'] + df_DataThemes['Sovereign_Score']) * df_DataThemes['Portfolio_Weight'] / df_DataThemes[df_DataThemes['Portfolio_Score_Contribution'] > 0].groupby('Pf')['Portfolio_Weight'].transform('sum')
    
    #Portfolio WACI Score
    Sector_Level2 = ['Industrial', 'Utility', 'Financial Institution', 'Quasi Sovereign']
    df_DataThemes['Portfolio_WACI_Corporates'] =  df_DataThemes.loc[df_DataThemes['Sector Level 2'].isin(Sector_Level2) & (df_DataThemes['CARBON_EMISSIONS_SCOPE_12_INTEN'] > 0), 'CARBON_EMISSIONS_SCOPE_12_INTEN'] * df_DataThemes.loc[df_DataThemes['Sector Level 2'].isin(Sector_Level2) & (df_DataThemes['CARBON_EMISSIONS_SCOPE_12_INTEN'] > 0), 'Portfolio_Weight'] / df_DataThemes[df_DataThemes['Sector Level 2'].isin(Sector_Level2) & (df_DataThemes['CARBON_EMISSIONS_SCOPE_12_INTEN'] > 0)].groupby('Pf')['Portfolio_Weight'].transform('sum')
    df_DataThemes['Portfolio_WACI_Corporates_Coverage'] = df_DataThemes.loc[df_DataThemes['Sector Level 2'].isin(Sector_Level2) & (df_DataThemes['CARBON_EMISSIONS_SCOPE_12_INTEN'] > 0), 'Portfolio_Weight'] / df_DataThemes[df_DataThemes['Sector Level 2'].isin(Sector_Level2)].groupby('Pf')['Portfolio_Weight'].transform('sum')     
    
    Sector_Level2 = ['Sovereign']
    df_DataThemes['Portfolio_WACI_Sovereign'] =  df_DataThemes.loc[df_DataThemes['Sector Level 2'].isin(Sector_Level2) & (df_DataThemes['CARBON_GOVERNMENT_GHG_INTENSITY_GDP_TONPERMN'] > 0), 'CARBON_GOVERNMENT_GHG_INTENSITY_GDP_TONPERMN'] * df_DataThemes.loc[df_DataThemes['Sector Level 2'].isin(Sector_Level2) & (df_DataThemes['CARBON_GOVERNMENT_GHG_INTENSITY_GDP_TONPERMN'] > 0), 'Portfolio_Weight'] / df_DataThemes[df_DataThemes['Sector Level 2'].isin(Sector_Level2) & (df_DataThemes['CARBON_GOVERNMENT_GHG_INTENSITY_GDP_TONPERMN'] > 0)].groupby('Pf')['Portfolio_Weight'].transform('sum')
    df_DataThemes['Portfolio_WACI_Sovereign_Coverage'] = df_DataThemes.loc[df_DataThemes['Sector Level 2'].isin(Sector_Level2) & (df_DataThemes['CARBON_GOVERNMENT_GHG_INTENSITY_GDP_TONPERMN'] > 0), 'Portfolio_Weight'] / df_DataThemes[df_DataThemes['Sector Level 2'].isin(Sector_Level2)].groupby('Pf')['Portfolio_Weight'].transform('sum')     

    #Assign Risk Score
    df_DataThemes['Risk_Score_Overall'] = 'Poor Risk Score'
    df_DataThemes.loc[((df_DataThemes['Corporate_Score'] <= 4) & (df_DataThemes['Corporate_Score'] > 2)) | ((df_DataThemes['Securitized_Score'] <= 4) & (df_DataThemes['Securitized_Score'] > 2)) | ((df_DataThemes['Sovereign_Score'] <= 4) & (df_DataThemes['Sovereign_Score'] > 2)) | ((df_DataThemes['Muni_Score'] <= 4) & (df_DataThemes['Muni_Score'] > 2)), 'Risk_Score_Overall'] = 'Average ESG Score'
    df_DataThemes.loc[((df_DataThemes['Corporate_Score'] <= 2) & (df_DataThemes['Corporate_Score'] >= 1)) | ((df_DataThemes['Securitized_Score'] <= 2) & (df_DataThemes['Securitized_Score'] >= 1)) | ((df_DataThemes['Sovereign_Score'] <= 2) & (df_DataThemes['Sovereign_Score'] >=1)) | ((df_DataThemes['Muni_Score'] <= 2) & (df_DataThemes['Muni_Score'] >= 1)), 'Risk_Score_Overall'] = 'Leading ESG Score'
    df_DataThemes.loc[(df_DataThemes['Corporate_Score'] < 1) & (df_DataThemes['Securitized_Score'] == 0) & (df_DataThemes['Sovereign_Score'] == 0) & (df_DataThemes['Muni_Score'] == 0), 'Risk_Score_Overall'] = 'Not Scored'

             
    #Delete Previous Output Sheet
    delete_unused_sheets(FilePathExport + 'ThemesList.xlsx')
    df_DataThemes = df_DataThemes.drop_duplicates(subset=['Pf', Identifier, 'Portfolio_Weight'], keep = 'first')
    
    write_to_excel(FilePathExport + 'ThemesList.xlsx', 'Subset',  df_DataThemes[['Pf',
    'ISSUER_NAME',
    'ISIN',
    'Portfolio_Weight',
    'Labeled ESG Type',
    'Sector Level 2',
    'BCLASS_Level4',
    'GICS_SUB_IND',
    'SClass_Level1',
    'SClass_Level2',
    'SClass_Level3',
    'SClass_Level4',
    'SClass_Level4-P',
    'SClass_Level5',
    'Review_Flag',
    'Review_Comments',
    'Securitized_Score',
    'Sovereign_Score',
    'Muni_Score',
    'ESRM_Score',
    'ESRM_Flagged',
    'ESRM_IsNAN',
    'PRIVACY_DATA_SEC_EXP_SCORE',
    'COMM_REL_RISK_EXP_SCORE',
    'LABOR_MGMT_EXP_SCORE',
    'WATER_STRESS_EXP_SCORE',
    'ENERGY_EFFICIENCY_EXP_SCORE',
    'BIODIV_LAND_USE_EXP_SCORE',
    'CONTR_SUPPLY_CHAIN_LABOR_N_TOTAL',
    'CUSTOMER_RELATIONS_SCORE',
    'PROD_SFTY_QUALITY_EXP_SCORE',
    'IVA_COMPANY_RATING',
    'OVERALL_FLAG',
    'UNGC_COMPLIANCE',
    'Governance_Score',
    'Governance_Flagged',
    'Governance_IsNAN',
    'CARBON_EMISSIONS_CDP_DISCLOSURE_Flag',
    'COMBINED_CEO_CHAIR_Flag',
    'CONTROLLING_SHAREHOLDER_Flag',
    'CROSS_SHAREHOLDINGS_Flag',
    'FEMALE_DIRECTORS_PCT_Flag',
    'INDEPENDENT_BOARD_MAJORITY_Flag',
    'NEGATIVE_DIRECTOR_VOTES_Flag',
    'PAY_LINKED_TO_SUSTAINABILITY_Flag',
    'POISON_PILL_Flag',
    'WOMEN_EXEC_MGMT_PCT_RECENT_Flag',
    'Transition_Score',
    'Transition_Sector',
    'Carbon Intensity (Scope 123)',
    'ClimateGHGReductionTargets_x',
    'HAS_SBTI_APPROVED_TARGET_x',
    'HAS_COMMITTED_TO_SBTI_TARGET_x',
    'CapEx',
    'Climate_Revenue',
    'RENEWENERGY_MSCI',
    'RENEWENERGY_ISS',
    'MOBILITY_MSCI',
    'MOBILITY_ISS',
    'CIRCULARITY_MSCI',
    'CIRCULARITY_ISS',
    'CCADAPT_MSCI',
    'CCADAPT_ISS',
    'BIODIVERSITY_MSCI',
    'BIODIVERSITY_ISS',
    'SMARTCITIES_MSCI',
    'SMARTCITIES_ISS',
    'EDU_MSCI',
    'EDU_ISS',
    'HEALTH_MSCI',
    'HEALTH_ISS',
    'SANITATION_MSCI',
    'SANITATION_ISS',
    'INCLUSION_MSCI',
    'INCLUSION_ISS',
    'NUTRITION_MSCI',
    'NUTRITION_ISS',
    'AFFORDABLE_MSCI',
    'AFFORDABLE_ISS']])
    
    write_to_excel(FilePathExport + 'ThemesList.xlsx', 'All_Files',  df_DataThemes)
    
    return df_DataThemes




#Initial program here...
#Load ISS Security Mapping and map to ISS SDG Datasets
#get parent
cwd = 'C:\\Users\\bastit\\Documents\\Risk_Score\\Input\\'
#Load ISS Security Mapping and map to ISS data
df_ISS_Security = pd.read_csv(cwd + 'Multi-Security_Standard_Issuers_20230403.csv')
df_ISS_SDG = pd.read_csv(cwd + 'TCW_SDG_Impact_Rating_Issuers_20230406.csv')
df_ISS_SDGA = pd.read_csv(cwd + 'TCW_SDGA_Issuers_20230406.csv')
df_ISS_Data_Transition = df_ISS_Security[['ISIN','Security ISIN']].merge(df_ISS_SDG[['ISIN', 'IssuerName', 'Ticker', 'CountryOfIncorporation', 'ESGRatingIndustry', 'ClimateGHGReductionTargets', 'BrownExpTotalCapExSharePercent', 'GreenExpTotalCapExSharePercent']], how = 'left', on= 'ISIN').dropna(subset = ['IssuerName'])    
df_ISS_Data_SDG_SDGA = df_ISS_SDGA.merge(df_ISS_SDG[['ISIN', 'IssuerName', 'Ticker', 'CountryOfIncorporation', 'ESGRatingIndustry', 'ClimateGHGReductionTargets', 'BrownExpTotalCapExSharePercent', 'GreenExpTotalCapExSharePercent']], 
                      how = 'left', on = ['ISIN', 'IssuerName', 'Ticker', 'CountryOfIncorporation', 'ESGRatingIndustry'])
df_ISS_Data_Themes = df_ISS_Security[['ISIN','Security ISIN']].merge(df_ISS_Data_SDG_SDGA, how = 'left', on= 'ISIN').dropna(subset = ['IssuerName'])


FileName = 'Portfolio_Data_043023.xlsx'
FilePathImport = 'C:\\Users\\bastit\\Documents\\Risk_Score\\Input\\'
FilePathExport  = 'C:\\Users\\abdars\\.spyder-py3\\Data Projects\\Export\\'
    
#Set Parameters
Q_Low = 0.10 #middle score
Q_High = 0.70 #middle score
High_Threshold = 1500


transition_score_calc(FileName, FilePathImport, FilePathExport, Q_Low, Q_High, High_Threshold, df_ISS_Data_Transition)
ESRM_Gov_score_calc(FileName, FilePathImport, FilePathExport)
df_Themes = themes_calc(FileName, FilePathImport, FilePathExport, df_ISS_Data_Themes)

